#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""easydsd v0.12 - DART 감사보고서 변환 도구
완전체: DSD↔Excel 무결점 엔진 + AI재무검증 + 전기금액검증 + DSD Diff비교 + 롤오버
"""
import os, re, sys, io, zipfile, threading, socket, time, json, base64
if sys.platform == 'win32':
    try:
        sys.stdout.reconfigure(encoding='utf-8', errors='replace')
        sys.stderr.reconfigure(encoding='utf-8', errors='replace')
    except Exception: pass

IS_FROZEN = getattr(sys, 'frozen', False) or hasattr(sys, '_MEIPASS')
BASE_DIR  = os.path.dirname(sys.executable if IS_FROZEN else os.path.abspath(__file__))

try:
    from flask import Flask, request, send_file, jsonify, render_template_string, Response
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    if IS_FROZEN: sys.exit(1)
    import subprocess
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'flask', 'openpyxl', '-q'])
    from flask import Flask, request, send_file, jsonify, render_template_string, Response
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

# AI SDK 선택적 로드 (없어도 기본 기능 동작)
try:
    import google.generativeai as genai
    GENAI_AVAILABLE = True
except ImportError:
    genai = None
    GENAI_AVAILABLE = False

try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    anthropic = None
    ANTHROPIC_AVAILABLE = False

# ── 1. 상수 정의 ──────────────────────────────────────────────────────────────
EDIT_COLOR   = 'FFF2CC'   # 일반 편집 셀 배경 (노랑)
HEADER_COLOR = 'DEEAF1'   # 헤더 행 배경 (파랑)
SUM_COLOR    = 'E0F7FA'   # 합계 행 배경 (하늘)
PARA_COLOR   = 'E8F5E9'   # 단락 배경 (연초록)
FMT_ACCOUNT  = '#,##0;(#,##0);"-"'
FMT_DECIMAL  = '#,##0.00'
FIN_TABLE_MAP = [
    (['재 무 상 태 표'], '🏦재무상태표'),
    (['포 괄 손 익 계 산 서'], '💹포괄손익계산서'),
    (['자 본 변 동 표'], '📈자본변동표'),
    (['현 금 흐 름 표'], '💰현금흐름표')
]
FIN_PREFIXES = ('🏦','💹','📈','💰')
SUM_KEYWORDS = ['합계','총계','합 계','총 계']

def fill(c): return PatternFill('solid', fgColor=c)
def fnt(color=None,bold=False,size=9): return Font(color=color if color else '000000', bold=bold, size=size)
def aln(h='left',v='center',wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

# ── 2. 원본 구조 파싱 & 식별 계 ──────────────────────────────────────────────
def clean_title(s):
    nl=' \n'
    return re.sub(r'\s+', ' ', s.replace('&amp;cr;', nl).replace('&cr;', nl).replace('&amp;', '&').replace('&lt;', '<').replace('&gt;', '>').replace('&quot;', '"')).strip()

def is_blank_title(s): return len(re.sub(r'[&;a-z]+','',s).strip())==0

def parse_cell(m):
    val=(re.sub(r'<[^>]+>', '', m.group(0)).replace('&amp;cr;','\n').replace('&amp;','&').replace('&lt;','<').replace('&gt;','>').replace('&quot;','"').replace('&cr;','\n').strip())
    cs = int(col_mat.group(1)) if (col_mat := re.search(r'COLSPAN=["\']?(\d+)["\']?', m.group(1), re.IGNORECASE)) else 1
    tag = re.match(r'<([A-Z]+)', m.group(0), re.IGNORECASE).group(1).upper()
    return dict(value=val, colspan=cs, tag=tag)

def parse_xml(xml):
    exts=re.findall(r'<EXTRACTION[^>]*ACODE="([^"]+)"[^>]*>([^<]+)</EXTRACTION>',xml)
    tables=[]
    for ti, tm in enumerate(re.finditer(r'<TABLE[^>]*>(?:(?!<TABLE).)*?</TABLE\s*>', xml, re.IGNORECASE | re.DOTALL)):
        ctx=xml[max(0, tm.start()-600):tm.start()]
        fin_label=next((lbl for kws,lbl in FIN_TABLE_MAP if any(kw in ctx or kw in tm.group(0) for kw in kws)),'')
        ctx_titles=[clean_title(t) for t in re.findall(r'<(?:TITLE|P)[^>]*>([^<]{3,80})</(?:TITLE|P)>', ctx) if not is_blank_title(t)]
        rows=[]
        for tr in re.finditer(r'<TR[^>]*>(.*?)</TR\s*>', tm.group(0), re.IGNORECASE | re.DOTALL):
            if cells :=[parse_cell(cm) for cm in re.finditer(r'<(?:TD|TH|TU|TE)([^>]*)>(.*?)</(?:TD|TH|TU|TE)\s*>', tr.group(1), re.IGNORECASE | re.DOTALL)]:
                rows.append(cells)
        tables.append(dict(idx=ti, fin_label=fin_label, ctx_title=ctx_titles[-1] if ctx_titles else '', rows=rows, start=tm.start()))
    return exts, tables

def is_num_or_decimal(val):
    v = (str(val).strip().replace(',','').replace('(','').replace(')','').replace('%','').replace('-','').replace(' ','').split('\n')[0])
    try: return float(v) or True
    except: return False

def _to_cell_value(v):
    s = str(v).strip(); neg = s.startswith('(') and s.endswith(')')
    if not s or s in ('-',''): return v
    if len(p:=s.split(','))>=2 and all(pt.strip().isdigit() and 1<=len(pt.strip())<=2 for pt in p): return v
    if not (cl:=s.replace(',','').replace('(','').replace(')','').replace(' ','')): return v
    try: return -float(cl) if neg else float(cl)
    except: return v

# ── 3. 숫자 정규화 (Excel→DSD 복원용) ───────────────────────────────────────
def is_note_ref(val):
    """주석 번호 참조 여부 (예: '1, 2, 3' 같은 짧은 숫자 나열)"""
    parts = val.strip().split(',')
    return (len(parts) >= 2 and all(p.strip().isdigit() and 1 <= len(p.strip()) <= 2 for p in parts))

def normalize_num(val):
    """Excel 셀 값을 DSD XML에 삽입 가능한 문자열로 변환
    - float('1234567.0') → '1,234,567'
    - 음수 float('-1234.0') → '(1,234)'
    - 텍스트는 그대로, HTML 특수문자 이스케이프
    """
    v = str(val).strip()
    if not v or v in ('None', ''): return ''
    if v == '-': return '-'
    if '\n' in v: return '&amp;cr;'.join(normalize_num(l) for l in v.split('\n'))
    if is_note_ref(v): return v
    neg = v.startswith('-') or (v.startswith('(') and v.endswith(')'))
    cl = v.replace(',','').replace('(','').replace(')','').replace('-','').replace(' ','')
    try:
        f = float(cl)
        fmt = f"{int(f):,}" if f == int(f) else f"{f:,.2f}"
        v = f"({fmt})" if neg else fmt
    except (ValueError, TypeError):
        pass
    return v.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('"','&quot;')

# ── 4. TD 정렬 보정 (밀린 행이 합계 TR 구조에 들어갈 때 CENTER→LEFT 방지) ──
def _is_numeric_val(v):
    cl = str(v).strip().replace(',','').replace('(','').replace(')','').replace('-','').replace(' ','')
    if not cl: return False
    try: float(cl); return True
    except: return False

def adjust_td_align(td_tag, value, is_header_row):
    """헤더 행이 아닌데 텍스트 값이 CENTER 정렬 TD에 들어올 경우 LEFT로 교정."""
    if is_header_row: return td_tag
    if _is_numeric_val(value): return td_tag
    return re.sub(r'\bALIGN="CENTER"', 'ALIGN="LEFT"', td_tag, flags=re.I)

# ── 5. [핵심] 색상 기반 Excel 테이블 블럭 추출기 ─────────────────────────────
def extract_excel_table_blocks(ws, EDIT_COLOR, HEADER_COLOR, SUM_COLOR):
    """(오프셋 붕괴 방어선 1조)
    색상 배경을 갖는 구역만 그룹핑! 테이블들이 동적으로 추가삭제 되더라도
    배열 묶음 자체를 매핑하므로 이탈 안 됨!"""
    blocks =[]
    current_block =[]
    empty_tides = 0
    for ri, rw in enumerate(ws.iter_rows(min_row=1), start=1):
        items_in_row =[]
        is_hit = False
        for ci, cell in enumerate(rw, start=1):
            if getattr(cell.fill, 'fgColor', None) and type(cell.fill.fgColor.rgb) == str:
                hc = cell.fill.fgColor.rgb.upper()
                if hc.endswith(EDIT_COLOR) or hc.endswith(SUM_COLOR) or hc.endswith(HEADER_COLOR):
                    items_in_row.append((ci, str(cell.value) if cell.value is not None else ''))
                    is_hit = True
        if is_hit:
            current_block.append((ri, items_in_row))
            empty_tides = 0
        else:
            empty_tides += 1
            if empty_tides >= 1 and current_block:
                blocks.append(current_block)
                current_block =[]
    if current_block: blocks.append(current_block)
    return blocks

# ── 6. [핵심 엔진] Excel → DSD 재조립기 ─────────────────────────────────────
def excel_to_dsd_bytes(xlsx_bytes):
    # ■ XLSX 바이너리 오염 및 보호 강제 차단용 사전 체크문
    if not xlsx_bytes or not xlsx_bytes.startswith(b'PK'):
        raise Exception("파일 구조 확인 실패 : '회사내 열람권한 보호 잠금 설정(마크애니 등등)'에 의해 엑셀 포맷을 읽을 수 없는 내부 파괴파일로 판단되었습니다. 정상(사내망 외 PC) 에서 잠김이 없는 형태로 저장바랍니다.")
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    if '_원본DSD_바이너리' not in wb.sheetnames:
        raise Exception("엑셀 안 DSD 변환 소스코드가 제거된 단순 조작파일이 되었습니다. 초기 1. 엑셀 풀기 작업을 받은 작업 폼 문서(.xlsx)의 시트를 그대로 복제 보존 바랍니다.")
    # 1. 기둥 XML 원천 부 추출
    try:
        ws_bin = wb['_원본DSD_바이너리']
        cx_core = base64.b64decode("".join([str(ws_bin.cell(r, 1).value) for r in range(1, ws_bin.max_row+1) if ws_bin.cell(r,1).value]))
        with zipfile.ZipFile(io.BytesIO(cx_core)) as zf:
            z_archv = {n: zf.read(n) for n in zf.namelist()}
        cx = z_archv['contents.xml'].decode('utf-8', errors='replace')
    except Exception as e:
        raise Exception("시스템 이식 중 데이터 역순 복사 불량 : 양식 소실 엑셀!")
    # 2. _원본XML 로 매핑 구성(절대 위치 무의미해진 상황. 이제 단순히 매치순서를 맞추는 보증 수표용)
    m_info = {}  # { 시트이름 : [ xml에 있었던 t_idx1, t_idx2 .. ] }
    if '_원본XML' in wb.sheetnames:
        for rw in wb['_원본XML'].iter_rows(min_row=4, values_only=True):
            if not rw or not rw[0] or str(rw[0])=='-': continue
            try:
                sn, ty = str(rw[0]).strip(), str(rw[5]).strip() if len(rw)>5 and rw[5] else 'TABLE'
                if ty == 'TABLE': m_info.setdefault(sn, []).append(int(rw[1]))
            except: pass

    # XML 정규 파편 파싱 위치 확립부 (껍데기 살리고 안전 보호하는 부분)
    table_positions = [(m.start(), m.end()) for m in re.finditer(r'<TABLE[^>]*>(?:(?!<TABLE).)*?</TABLE\s*>', cx, re.IGNORECASE | re.DOTALL)]
    p_ins_block =[]
    # 3. 데이터 결합 동적엔진(Dynamic Align Layered Loop) 가동
    for s_idx, sn in enumerate(wb.sheetnames):
        if sn in ('📋사용안내', '_원본XML', '_원본DSD_바이너리', '🤖AI검증결과'): continue
        if not (mapped_tidx := m_info.get(sn)): continue # 해당 시트에 매칭할 테이블정보 리스트!
        ws_curr = wb[sn]
        t_chunks = extract_excel_table_blocks(ws_curr, EDIT_COLOR, HEADER_COLOR, SUM_COLOR)

        for k_idx, target_xml_tbl_idx in enumerate(mapped_tidx):
            if target_xml_tbl_idx >= len(table_positions): continue
            if k_idx >= len(t_chunks): break  # 혹시 테이블 통채로 유저가 날렸다면 스킵.
            # 해당 조각 구역의 2D 데이터 매핑 준비
            curr_xcl_block = t_chunks[k_idx]
            b_data_map = {}
            for loc_tr_idx, (real_exl_ri, lst_val_by_cols) in enumerate(curr_xcl_block):
                for e_ci, t_val in lst_val_by_cols:
                    b_data_map[(loc_tr_idx, e_ci - 1)] = t_val   # xcol 보정(0_idx 화)
            ts, te = table_positions[target_xml_tbl_idx]; tbl_cx_src = cx[ts:te]
            trs = list(re.finditer(r'(<TR[^>]*>)(.*?)(</TR\s*>)', tbl_cx_src, re.IGNORECASE | re.DOTALL))
            if not trs: continue
            c_parts, p_tail =[], 0
            n_tgt_row_num = max(len(trs), len(curr_xcl_block))
            # 클론 행 템플릿: trs[-1](합계 행)이 아닌 일반 데이터 행(두 번째 TR) 사용 → CENTER 정렬 오염 방지
            clone_tpl = trs[min(1, len(trs)-1)]
            # [마이크로 TR 리턴 치환 공정시작!] -- 행밀림 에러/가운데 정렬깨짐, XML 태그 닫기 파편증상 ALL DELETE --
            for n_ti in range(n_tgt_row_num):
                cloned = (n_ti >= len(trs))
                hit_rm = clone_tpl if cloned else trs[n_ti]
                if not cloned: c_parts.append(tbl_cx_src[p_tail:hit_rm.start()])

                openTag, midVal, closeTag = hit_rm.group(1), hit_rm.group(2), hit_rm.group(3)
                col_offset_head = 0
                v_body =[]
                c_td_last = 0

                for inner_TD_m in re.finditer(r'(<(?:TD|TH|TU|TE)[^>]*>)(.*?)(</(?:TD|TH|TU|TE)\s*>)', midVal, re.IGNORECASE | re.DOTALL):
                    v_body.append(midVal[c_td_last : inner_TD_m.start()])
                    TD_o, content_chk, TD_c = inner_TD_m.group(1), inner_TD_m.group(2), inner_TD_m.group(3)

                    cSP = int(cf_.group(1)) if (cf_:=re.search(r'COLSPAN=["\']?(\d+)["\']?', TD_o, re.I)) else 1
                    lookup_pt = (n_ti, col_offset_head)

                    wrap_g = re.match(r'^(\s*<(?:P|SPAN|DIV|FONT|TITLE)[^>]*>)(.*?)(</(?:P|SPAN|DIV|FONT|TITLE)>\s*)$', content_chk, re.IGNORECASE | re.DOTALL)

                    if lookup_pt in b_data_map: # 값 유입. 삽입 (더욱 간소화 정제하여 HTML 결합에 의한 치환사고 없게 함)
                        clean_target_text = b_data_map[lookup_pt]
                        cT = normalize_num(clean_target_text)
                        sIn = f"{wrap_g.group(1)}{cT}{wrap_g.group(3)}" if wrap_g else cT
                        # 밀린 행이 합계 TR 구조에 들어올 때 CENTER 정렬 교정 (헤더 행 n_ti=0 제외)
                        fixed_TD_o = adjust_td_align(TD_o, clean_target_text, n_ti == 0)
                        v_body.append(fixed_TD_o + sIn + TD_c)
                    else:
                        if cloned:
                            cE = f"{wrap_g.group(1)}&amp;nbsp;{wrap_g.group(3)}" if wrap_g else "&amp;nbsp;"
                            v_body.append(TD_o + cE + TD_c)
                        else:
                            v_body.append(inner_TD_m.group(0))
                    c_td_last = inner_TD_m.end(); col_offset_head += cSP

                v_body.append(midVal[c_td_last:])
                c_parts.append(openTag + "".join(v_body) + closeTag)
                if not cloned: p_tail = hit_rm.end()
            # TBODY, THEAD 최외곽 XML구조가 끝맺도록 결합! (가운데정렬 파쇄나 에디터 인식 불가 회피완성됨)
            c_parts.append(tbl_cx_src[p_tail:])
            p_ins_block.append((ts, te, "".join(c_parts)))

    # 전체 오프셋 인덱스를 어그러뜨리지 않기위하여 맨 끝 테이블서부터 적용 (1차, DART 역호환 원칙 채용)
    rslt_cx = cx
    for bs, be, htmls in sorted(p_ins_block, key=lambda i:-i[0]): rslt_cx = rslt_cx[:bs] + htmls + rslt_cx[be:]

    buff = io.BytesIO()
    with zipfile.ZipFile(buff, 'w', zipfile.ZIP_DEFLATED) as zh:
        for oriN, dBit in z_archv.items():
            if oriN.endswith(('.png','.jpeg','.jpg','.gif')): continue
            zh.writestr(oriN, rslt_cx.encode('utf-8') if oriN=='contents.xml' else dBit)
    return buff.getvalue()

# ── 7. [핵심 엔진] DSD → Excel 빌더 ─────────────────────────────────────────
def dsd_to_excel_bytes(dsd_bytes, do_period_change=False, period_params=None):
    with zipfile.ZipFile(io.BytesIO(dsd_bytes)) as zf:
        xml = zf.read('contents.xml').decode('utf-8', errors='replace')
    exts, tables = parse_xml(xml)

    wb = openpyxl.Workbook(); ws0 = wb.active; ws0.title = '📋사용안내'; ws0.sheet_view.showGridLines=False
    for ri,(tx,bl,fg,bg,sz) in enumerate([
        ('✅ DART 문서 완벽 템플릿(결속방어형 탑재 시스템)',1,'1F4E79','FFFFFF',13),('',0,'000000','FFFFFF',8),
        ('[주의 1]: 가장 큰 문제는 행 자체를 날려 빈 줄 처리 후 아래 테이블들의 포맷(결속)이 손상된 탓 입니다!', 1,'FFFFFF','B71C1C',10),
        ('[필독!!] 셀을 줄째로 [삭제]하거나 빈줄[삽입]하면 인식이 고착됩니다! 줄째 카피 하여[복사된 셀 삽입], 혹은 [해당값 비우기 및 - 긋기]만 유지하세요!!', 1,'FFFFFF','E65100',11),
    ], 1):
        c=ws0.cell(ri, 1, tx); c.font=fnt(color=fg, bold=bl, size=sz)
        ws0.column_dimensions['A'].width=100
        if bg and bg != 'FFFFFF': c.fill=fill(bg)
    map_data =[]
    ftbs = [t for t in tables if t.get('fin_label')]
    for title, tbl, mx in ([(x['fin_label'], [x], False) for x in ftbs]):
        bn=re.sub(r'[\\/*?:\[\]]', '', title).strip(); sname=bn[:31]; ws=wb.create_sheet(sname)
        map_data.append((ws.title, tbl[0]['idx'], 0, 'TABLE', 1, 'TABLE'))
        erow=1; erow_start=1  # head 판단은 시작행 기준 고정값 사용 (rx==erow라 항상True되던 버그 수정)
        for rx, rw in enumerate(tbl[0]['rows'], erow):
            head, sm, cid = (rx < erow_start+2), any(kw in "".join([z['value'] for z in rw if isinstance(z,dict)]).replace(' ','') for kw in SUM_KEYWORDS), 1
            for cn in rw:
                if not isinstance(cn, dict): continue
                tc = ws.cell(erow, cid, _to_cell_value(cn['value']))
                if head: tc.fill, tc.alignment, tc.font = fill(HEADER_COLOR), aln('center','center',True), fnt(bold=True)
                else: tc.fill = fill(SUM_COLOR) if sm else fill(EDIT_COLOR)
                if is_num_or_decimal(cn['value']) and not head: tc.number_format, tc.alignment = FMT_ACCOUNT if '.' not in cn['value'] else FMT_DECIMAL, aln('right','center')
                cid += cn.get('colspan', 1)
            erow+=1

    b_rest =[r for r in tables if not r.get('fin_label')]
    cst=1
    for rest in b_rest:
        bx = (rest.get('ctx_title','').replace('/','').replace('[','')[:10])
        sn = f"📝_{bx}_{cst}"[:31]; cst+=1; ws = wb.create_sheet(sn); er=1; er_start=1
        map_data.append((ws.title, rest['idx'], 0, 'TABLE', er, 'TABLE'))
        for rI, rt in enumerate(rest['rows'], er):
            HD, sUM, d_id = (rI < er_start+2), any(h in "".join([y['value'] for y in rt if type(y)==dict]).replace(' ','') for h in SUM_KEYWORDS), 1
            for CI in rt:
                if not isinstance(CI,dict): continue
                bc = ws.cell(er, d_id, _to_cell_value(CI['value']))
                if HD: bc.fill, bc.font, bc.alignment = fill(HEADER_COLOR), fnt(bold=True), aln('center','center',True)
                else: bc.fill = fill(SUM_COLOR) if sUM else fill(EDIT_COLOR)
                if is_num_or_decimal(CI['value']) and not HD: bc.number_format, bc.alignment = FMT_ACCOUNT if '.' not in CI['value'] else FMT_DECIMAL, aln('right','center')
                d_id += CI.get('colspan',1)
            er+=1
    wm = wb.create_sheet('_원본XML'); wm.sheet_state='hidden'; wm.append(['SN','is','ie','DT','er','rt']); wm.append(['-']*6); wm.append(['-']*6); wm.append(['-']*6)
    for dm in map_data: wm.append(list(dm))
    wB = wb.create_sheet('_원본DSD_바이너리'); wB.sheet_state='hidden'
    bfB = base64.b64encode(dsd_bytes).decode('utf-8')
    for rx, ci in enumerate(range(0, len(bfB), 30000), 1): wB.cell(rx, 1, bfB[ci:ci+30000])
    buf = io.BytesIO(); wb.save(buf); xlsx_out = buf.getvalue()

    # ── 롤오버 후처리: do_period_change=True 시 당기→전기 이관 적용 ──────────
    if do_period_change:
        xlsx_out = apply_rollover(xlsx_out, period_params or {})

    return xlsx_out

# ── 8. 롤오버 후처리 엔진 ─────────────────────────────────────────────────────
def apply_rollover(xlsx_bytes, params):
    """당기 숫자를 전기 칸으로 밀고 헤더의 기수/연도를 자동 변경.
    - HEADER_COLOR 행에서 당기/전기 컬럼 위치 파악
    - 당기 값 → 전기 슬롯으로 복사, 당기 슬롯은 비움
    - 기수 문자열 자동 증가 (제 N 기 → 제 N+1 기)
    - 연도 문자열 자동 증가 (params['year_offset'] 만큼)
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))

    for sn in wb.sheetnames:
        # 시스템 시트 스킵
        if sn in ('📋사용안내', '_원본XML', '_원본DSD_바이너리', '🤖AI검증결과'):
            continue
        ws = wb[sn]

        # 헤더 행에서 당기/전기 컬럼 위치 파악
        # { 헤더행번호: {'당기': [col_list], '전기': [col_list]} }
        header_col_map = {}
        for ri, row in enumerate(ws.iter_rows(), start=1):
            for ci, cell in enumerate(row, start=1):
                if not (getattr(cell.fill, 'fgColor', None) and
                        type(cell.fill.fgColor.rgb) == str and
                        cell.fill.fgColor.rgb.upper().endswith(HEADER_COLOR)):
                    continue
                val = str(cell.value or '').strip()
                if re.search(r'당\s*기', val):
                    header_col_map.setdefault(ri, {'당기': [], '전기': []})['당기'].append(ci)
                elif re.search(r'전\s*기', val):
                    header_col_map.setdefault(ri, {'당기': [], '전기': []})['전기'].append(ci)

        if not header_col_map:
            continue

        # 첫 번째 헤더 행 기준으로 당기/전기 컬럼 결정
        for h_row, col_info in sorted(header_col_map.items()):
            dk_cols = col_info.get('당기', [])
            jk_cols = col_info.get('전기', [])
            if not dk_cols or not jk_cols:
                continue

            # 데이터 행: 헤더 다음 행부터 끝까지
            for data_ri in range(h_row + 1, ws.max_row + 1):
                for pair_idx, dk_col in enumerate(dk_cols):
                    if pair_idx < len(jk_cols):
                        jk_col = jk_cols[pair_idx]
                        dk_cell = ws.cell(data_ri, dk_col)
                        jk_cell = ws.cell(data_ri, jk_col)
                        # 당기 값을 전기로 복사
                        if dk_cell.value is not None:
                            jk_cell.value = dk_cell.value
                            jk_cell.number_format = dk_cell.number_format
                            jk_cell.alignment = dk_cell.alignment
                        # 당기 셀 비우기
                        dk_cell.value = None

            # 헤더 기수/날짜 자동 갱신
            _update_period_headers(ws, h_row, params)
            break  # 시트당 첫 번째 헤더 세트만 처리

    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def _update_period_headers(ws, h_row, params):
    """헤더 행의 기수(제 N 기) 및 연도 문자열을 자동 증가 변환."""
    year_offset = int(params.get('year_offset', 1))
    for ci in range(1, ws.max_column + 1):
        cell = ws.cell(h_row, ci)
        val = str(cell.value or '')
        if not val.strip():
            continue
        # 기수 패턴: '제 N 기' 또는 '제N기' → N+1로 증가
        def increment_period(m):
            n = int(m.group(1))
            return m.group(0).replace(str(n), str(n + 1), 1)
        new_val = re.sub(r'제\s*(\d+)\s*기', increment_period, val)
        # 연도 4자리 패턴 자동 증가
        if year_offset != 0:
            new_val = re.sub(r'(\d{4})', lambda m: str(int(m.group(1)) + year_offset), new_val)
        if new_val != val:
            cell.value = new_val

# ── 9. AI 검증용 Excel 분석기 ─────────────────────────────────────────────
def _safe_float(v):
    """문자/숫자/괄호 음수 형태를 안전하게 float로 변환."""
    if v is None:
        return None
    s = str(v).strip()
    if not s or s in ('-', 'None', 'nan'):
        return None
    neg = s.startswith('(') and s.endswith(')')
    s = s.replace(',', '').replace('(', '').replace(')', '').replace('%', '').replace(' ', '')
    try:
        n = float(s)
        return -n if neg else n
    except Exception:
        return None

def _iter_visible_sheets(wb):
    """사용자 편집 대상 시트만 순회."""
    for sn in wb.sheetnames:
        if sn in ('📋사용안내', '_원본XML', '_원본DSD_바이너리', '🤖AI검증결과'):
            continue
        ws = wb[sn]
        if getattr(ws, 'sheet_state', 'visible') != 'visible':
            continue
        yield ws

def _sheet_rows_as_matrix(ws):
    """시트의 비어있지 않은 행만 2차원 배열로 수집."""
    rows = []
    for ri in range(1, ws.max_row + 1):
        vals = [ws.cell(ri, ci).value for ci in range(1, ws.max_column + 1)]
        if any(v not in (None, '') for v in vals):
            rows.append(vals)
    return rows

def _find_balance_numbers(rows):
    """재무상태표에서 자산/부채/자본 총계 후보를 탐지."""
    assets = liabilities = equity = None
    for row in rows:
        texts = [str(v).strip() for v in row if v not in (None, '')]
        if not texts:
            continue
        first = texts[0].replace(' ', '')
        nums = [_safe_float(v) for v in row]
        num_vals = [n for n in nums if n is not None]
        if not num_vals:
            continue
        last_num = num_vals[-1]
        if any(k in first for k in ['자산총계', '자산계', '자산총합']):
            assets = last_num
        elif any(k in first for k in ['부채총계', '부채계', '부채총합']):
            liabilities = last_num
        elif any(k in first for k in ['자본총계', '자본계', '자본총합']):
            equity = last_num
    return assets, liabilities, equity

def _collect_numeric_profile(ws):
    """시트 전체 숫자 분포와 단위 이상치 후보를 탐지."""
    nums = []
    suspicious = []
    for ri in range(1, ws.max_row + 1):
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(ri, ci)
            n = _safe_float(cell.value)
            if n is None:
                continue
            nums.append(abs(n))
            if abs(n) != 0:
                digits = len(str(int(abs(n)))) if abs(n) >= 1 else 0
                suspicious.append({'row': ri, 'col': ci, 'value': str(cell.value), 'digits': digits})
    return nums, suspicious

def analyze_excel_financials(xlsx_bytes):
    """업로드된 Excel을 Python으로 정적 검증.
    - 대차평균
    - 합계/총계 라인 존재 여부
    - 숫자 자릿수 분포 기반 단위 이상 여부
    """
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    result = {
        'summary': {'sheet_count': 0, 'numeric_cell_count': 0, 'warning_count': 0, 'error_count': 0},
        'balance_checks': [],
        'unit_anomalies': [],
        'sum_row_checks': [],
        'sheets': []
    }

    for ws in _iter_visible_sheets(wb):
        rows = _sheet_rows_as_matrix(ws)
        nums, suspicious = _collect_numeric_profile(ws)
        result['summary']['sheet_count'] += 1
        result['summary']['numeric_cell_count'] += len(nums)

        sheet_report = {'sheet': ws.title, 'numeric_count': len(nums), 'sum_keywords_found': 0, 'warnings': []}

        # 1) 대차평균 검사
        assets, liabilities, equity = _find_balance_numbers(rows)
        if assets is not None and liabilities is not None and equity is not None:
            diff = assets - (liabilities + equity)
            ok = abs(diff) < 1
            result['balance_checks'].append({
                'sheet': ws.title,
                'assets': assets,
                'liabilities': liabilities,
                'equity': equity,
                'difference': diff,
                'status': 'OK' if ok else 'ERROR'
            })
            if ok:
                sheet_report['warnings'].append('자산총계 = 부채총계 + 자본총계 일치')
            else:
                sheet_report['warnings'].append(f'대차 불일치 감지: 차이 {diff:,.0f}')
                result['summary']['error_count'] += 1

        # 2) 합계/총계 라인 존재 여부
        for row in rows:
            first = str(row[0]).strip().replace(' ', '') if row and row[0] not in (None, '') else ''
            if any(k.replace(' ', '') in first for k in SUM_KEYWORDS):
                sheet_report['sum_keywords_found'] += 1
        result['sum_row_checks'].append({
            'sheet': ws.title,
            'sum_rows_found': sheet_report['sum_keywords_found'],
            'status': 'OK' if sheet_report['sum_keywords_found'] > 0 else 'WARN'
        })
        if sheet_report['sum_keywords_found'] == 0:
            sheet_report['warnings'].append('합계/총계 행이 탐지되지 않았습니다.')
            result['summary']['warning_count'] += 1

        # 3) 숫자 자릿수 편차 기반 단위 이상치
        if nums:
            positive = [n for n in nums if n > 0]
            if positive:
                digit_lengths = [len(str(int(n))) if n >= 1 else 0 for n in positive]
                max_digits = max(digit_lengths)
                min_digits = min(digit_lengths)
                if max_digits - min_digits >= 4:
                    top_suspicious = sorted(
                        [x for x in suspicious if x['digits'] in (max_digits, min_digits)],
                        key=lambda x: (-x['digits'], x['row'], x['col'])
                    )[:8]
                    result['unit_anomalies'].append({
                        'sheet': ws.title,
                        'digit_span': max_digits - min_digits,
                        'examples': top_suspicious,
                        'status': 'WARN'
                    })
                    sheet_report['warnings'].append(
                        f'숫자 자릿수 편차가 큽니다. 단위 혼입 가능성 확인 필요 (최소 {min_digits}자리 / 최대 {max_digits}자리)'
                    )
                    result['summary']['warning_count'] += 1

        result['sheets'].append(sheet_report)

    return result

def build_ai_prompt_from_excel_summary(summary):
    """Python 검증 결과를 AI에 넘길 프롬프트로 직렬화."""
    return f"""다음은 EasyDSD가 업로드된 Excel 재무제표를 Python으로 1차 검증한 결과입니다.
이 결과를 바탕으로 한국어로 재무제표 검증 의견을 작성하세요.

검토 지시:
1. 대차평균(자산총계 = 부채총계 + 자본총계) 결과를 검토
2. 합계/총계 행 탐지 결과를 검토
3. 숫자 자릿수 편차를 바탕으로 단위 이상 가능성을 판단
4. 사용자가 다시 확인해야 할 항목을 우선순위로 정리
5. 결과를 [✅통과 / ⚠️주의 / ❌오류] 형식으로 설명

Python 검증 JSON:
{json.dumps(summary, ensure_ascii=False, indent=2)}
"""

def run_ai_validation(summary, provider, api_key, model):
    """Gemini 또는 Claude를 호출하여 Python 검증 결과를 사람이 읽기 쉬운 보고서로 재서술."""
    if not api_key:
        return 'API Key가 없어 Python 검증만 수행했습니다.'
    prompt = build_ai_prompt_from_excel_summary(summary)
    provider = (provider or 'gemini').strip().lower()

    if provider == 'gemini':
        if not GENAI_AVAILABLE:
            raise Exception('google-generativeai 패키지가 없어 Gemini를 호출할 수 없습니다.')
        genai.configure(api_key=api_key)
        model_name = model or 'gemini-1.5-flash'
        resp = genai.GenerativeModel(model_name).generate_content(prompt)
        return getattr(resp, 'text', str(resp))

    if provider in ('anthropic', 'claude'):
        if not ANTHROPIC_AVAILABLE:
            raise Exception('anthropic 패키지가 없어 Claude를 호출할 수 없습니다.')
        model_name = model or 'claude-3-5-sonnet-latest'
        client = anthropic.Anthropic(api_key=api_key)
        resp = client.messages.create(
            model=model_name,
            max_tokens=1800,
            messages=[{'role': 'user', 'content': prompt}]
        )
        parts = []
        for block in getattr(resp, 'content', []):
            txt = getattr(block, 'text', None)
            if txt:
                parts.append(txt)
        return '\n'.join(parts).strip() or str(resp)

    raise Exception('지원하지 않는 AI 제공자입니다. gemini 또는 anthropic를 사용하세요.')

# ── 10. DSD에서 재무 텍스트 추출 (보조 도구) ─────────────────────────────────
def extract_financial_text_from_dsd(dsd_bytes):
    """DSD 파일에서 재무제표 테이블 텍스트를 추출."""
    with zipfile.ZipFile(io.BytesIO(dsd_bytes)) as zf:
        xml = zf.read('contents.xml').decode('utf-8', errors='replace')
    _, tables = parse_xml(xml)
    lines = []
    for tbl in tables:
        if not tbl.get('fin_label'):
            continue
        lines.append(f"\n=== {tbl['fin_label']} ===")
        for row in tbl['rows']:
            row_texts = [cell['value'] for cell in row if isinstance(cell, dict)]
            lines.append(' | '.join(row_texts))
    return '\n'.join(lines)

# ── 11. 전기금액 검증 (두 DSD 비교) ─────────────────────────────────────────
def verify_prior_period(prev_dsd_bytes, curr_dsd_bytes):
    """전년도 DSD의 당기 금액 vs 당해연도 DSD의 전기 금액 비교.
    롤오버 후 이관이 정확한지 항목별로 검증."""
    def extract_period_values(dsd_bytes, col_keyword):
        """특정 기간(당기/전기) 컬럼의 항목명:금액 딕셔너리 반환."""
        with zipfile.ZipFile(io.BytesIO(dsd_bytes)) as zf:
            xml = zf.read('contents.xml').decode('utf-8', errors='replace')
        _, tables = parse_xml(xml)
        result = {}  # { 재무제표라벨: { 항목명: 금액 } }
        for tbl in tables:
            if not tbl.get('fin_label'):
                continue
            label = tbl['fin_label']
            rows = tbl['rows']
            if not rows:
                continue
            # 헤더 행에서 대상 컬럼 위치 파악 (colspan 누적으로 실제 열 인덱스 계산)
            header_row = rows[0]
            target_cols = []
            col_cursor = 0
            for cell in header_row:
                if isinstance(cell, dict):
                    if re.search(col_keyword, cell['value']):
                        target_cols.append(col_cursor)
                    col_cursor += cell.get('colspan', 1)
            if not target_cols:
                continue
            # 데이터 행에서 항목명과 금액 추출
            tbl_data = {}
            for row in rows[1:]:
                if not row:
                    continue
                item_name = row[0]['value'].strip() if isinstance(row[0], dict) else ''
                if not item_name or item_name in ['-', '']:
                    continue
                col_cursor2 = 0
                for cell in row:
                    if not isinstance(cell, dict):
                        continue
                    if col_cursor2 in target_cols:
                        val = cell['value'].strip()
                        if val and val not in ['-', '']:
                            tbl_data[item_name] = val
                            break
                    col_cursor2 += cell.get('colspan', 1)
            if tbl_data:
                result[label] = tbl_data
        return result

    prev_current = extract_period_values(prev_dsd_bytes, r'당\s*기')  # 전년도 당기
    curr_prior   = extract_period_values(curr_dsd_bytes,  r'전\s*기')  # 당해연도 전기

    total_ok = total_mismatch = total_missing = 0
    report = []
    all_labels = set(list(prev_current.keys()) + list(curr_prior.keys()))

    for label in sorted(all_labels):
        prev_data = prev_current.get(label, {})
        curr_data = curr_prior.get(label, {})
        all_items = set(list(prev_data.keys()) + list(curr_data.keys()))
        table_rows = []
        for item in sorted(all_items):
            pv = prev_data.get(item, '(없음)')
            cv = curr_data.get(item, '(없음)')
            # 숫자 정규화 후 비교
            def _norm(v):
                v = str(v).strip().replace(',','').replace('(', '-').replace(')','')
                try: return float(v)
                except: return v
            if pv == '(없음)' or cv == '(없음)':
                status = '⚠️누락'; total_missing += 1
            elif _norm(pv) == _norm(cv):
                status = '✅일치'; total_ok += 1
            else:
                status = '❌불일치'; total_mismatch += 1
            table_rows.append({'item': item, 'prev_val': pv, 'curr_val': cv, 'status': status})
        report.append({'label': label, 'rows': table_rows})

    return {
        'summary': {'ok': total_ok, 'mismatch': total_mismatch, 'missing': total_missing},
        'tables': report
    }

# ── 12. DSD Diff 비교 분석 ────────────────────────────────────────────────────
def compare_dsd_files(before_dsd_bytes, after_dsd_bytes):
    """수정 전/후 DSD의 XML 노드를 전수 비교하여 변동된 셀만 추출.
    테이블 인덱스, 행/열 인덱스, 변경 전/후 값을 반환."""
    def get_all_cells(dsd_bytes):
        with zipfile.ZipFile(io.BytesIO(dsd_bytes)) as zf:
            xml = zf.read('contents.xml').decode('utf-8', errors='replace')
        _, tables = parse_xml(xml)
        cells = {}  # { (table_idx, label, row_idx, col_idx): value }
        for tbl in tables:
            t_idx = tbl['idx']
            label = tbl.get('fin_label') or tbl.get('ctx_title', f'테이블{t_idx}')
            for ri, row in enumerate(tbl['rows']):
                col_cursor = 0
                for cell in row:
                    if isinstance(cell, dict):
                        cells[(t_idx, label, ri, col_cursor)] = cell['value']
                        col_cursor += cell.get('colspan', 1)
        return cells

    before_cells = get_all_cells(before_dsd_bytes)
    after_cells  = get_all_cells(after_dsd_bytes)
    all_keys = set(list(before_cells.keys()) + list(after_cells.keys()))

    diffs = []; added = []; removed = []
    for key in sorted(all_keys, key=lambda k: (k[0], k[2], k[3])):
        t_idx, label, ri, ci = key
        bval = before_cells.get(key)
        aval = after_cells.get(key)
        if bval is None and aval is not None:
            added.append({'table': label, 't_idx': t_idx, 'row': ri, 'col': ci, 'after': aval})
        elif bval is not None and aval is None:
            removed.append({'table': label, 't_idx': t_idx, 'row': ri, 'col': ci, 'before': bval})
        elif bval != aval:
            diffs.append({'table': label, 't_idx': t_idx, 'row': ri, 'col': ci, 'before': bval, 'after': aval})

    return {
        'changed': diffs, 'added': added, 'removed': removed,
        'summary': {'changed_count': len(diffs), 'added_count': len(added), 'removed_count': len(removed)}
    }


# ── 13. 추가 후처리/점검/정리 기능 ───────────────────────────────────────────
from copy import copy
from datetime import datetime

def friendly_error_message(exc, stage='작업'):
    """사용자에게 보여줄 친화적 오류 문구."""
    msg = str(exc) or exc.__class__.__name__
    raw = msg.lower()
    if 'styleproxy' in raw:
        return f'{stage} 중 엑셀 스타일 처리 단계에서 오류가 발생했습니다. 핵심 변환 엔진이 아니라 후처리 단계 문제입니다. 옵션을 줄여 다시 시도해 주세요.'
    if 'xlsx file format cannot be determined' in raw or 'zip' in raw:
        return f'{stage}할 파일 형식을 읽지 못했습니다. 실제 .xlsx / .dsd 파일인지 확인해 주세요.'
    if '_원본dsd_바이너리' in msg or '_원본xml' in msg:
        return f'{stage}에 필요한 숨김 시트가 손상되었거나 삭제되었습니다. EasyDSD가 생성한 원본 작업양식을 사용해 주세요.'
    if 'permission' in raw:
        return f'{stage} 중 파일 접근 권한 문제로 실패했습니다. 파일을 닫고 다시 시도해 주세요.'
    if 'api key' in raw:
        return msg
    return f'{stage} 중 오류가 발생했습니다: {msg}'

def _copy_cell_style(src, dst):
    """StyleProxy 오류를 피하기 위해 속성을 안전 복사."""
    if src.has_style:
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.protection = copy(src.protection)
        dst.number_format = src.number_format

def _read_original_dsd_bytes_from_wb(wb):
    if '_원본DSD_바이너리' not in wb.sheetnames:
        return None
    ws = wb['_원본DSD_바이너리']
    buf = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v:
            buf.append(str(v))
    if not buf:
        return None
    return base64.b64decode(''.join(buf))

def detect_note_anchor_conflicts(tables):
    """주석 번호 충돌/중복/누락 간단 탐지."""
    anchors = []
    seen = {}
    duplicates = []
    missing_gaps = []
    PATS = [
        r'^주석\s*(\d{1,2})\s*[.\-·]\s*(.{2,30})',
        r'^(\d{1,2})\s*\.\s*([^\d\(].{1,30})',
        r'^\((\d{1,2})\)\s*([^\d].{1,30})',
    ]
    for tbl in tables:
        ctx = (tbl.get('ctx_title') or '').strip()
        if not ctx:
            continue
        for pat in PATS:
            m = re.match(pat, ctx)
            if m:
                n = int(m.group(1))
                title = clean_title(m.group(2))[:20]
                anchors.append({'note_num': n, 'title': title, 'table_idx': tbl['idx']})
                if n in seen:
                    duplicates.append({'note_num': n, 'first_title': seen[n], 'dup_title': title, 'table_idx': tbl['idx']})
                else:
                    seen[n] = title
                break
    nums = sorted({a['note_num'] for a in anchors})
    if nums:
        for i in range(nums[0], nums[-1]):
            if i not in nums:
                missing_gaps.append(i)
    return {'anchors': anchors, 'duplicates': duplicates, 'missing_numbers': missing_gaps}

def precheck_dsd_bytes(dsd_bytes):
    with zipfile.ZipFile(io.BytesIO(dsd_bytes)) as zf:
        xml = zf.read('contents.xml').decode('utf-8', errors='replace')
    exts, tables = parse_xml(xml)
    fin = [t for t in tables if t.get('fin_label')]
    notes = [t for t in tables if not t.get('fin_label')]
    conflict = detect_note_anchor_conflicts(tables)
    warnings = []
    if not fin:
        warnings.append('재무제표 TABLE을 찾지 못했습니다.')
    if conflict['duplicates']:
        warnings.append(f'주석 번호 중복 의심 {len(conflict["duplicates"])}건')
    if conflict['missing_numbers']:
        warnings.append(f'주석 번호 공백 의심: {conflict["missing_numbers"][:10]}')
    return {
        'table_count': len(tables),
        'financial_table_count': len(fin),
        'note_table_count': len(notes),
        'extraction_count': len(exts),
        'note_conflicts': conflict,
        'warnings': warnings,
    }

def precheck_excel_bytes(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    issues = []
    repairable = []
    fatal = []
    hidden_ok = True
    for req in ('_원본XML', '_원본DSD_바이너리'):
        if req not in wb.sheetnames:
            fatal.append(f'필수 숨김 시트 누락: {req}')
            hidden_ok = False
        elif wb[req].sheet_state != 'hidden':
            issues.append(f'{req} 시트가 숨김 해제되어 있습니다.')
            repairable.append(f'{req} 시트를 다시 hidden 처리할 수 있습니다.')

    visible_targets = [sn for sn in wb.sheetnames if sn not in ('📋사용안내','_원본XML','_원본DSD_바이너리','🤖AI검증결과')]
    if not visible_targets:
        fatal.append('편집 대상 시트가 없습니다.')

    if '_원본XML' in wb.sheetnames:
        ws = wb['_원본XML']
        mapped = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or not row[0] or row[0] == '-':
                continue
            mapped.append(str(row[0]).strip())
        missing = sorted(set(mapped) - set(wb.sheetnames))
        if missing:
            fatal.append(f'_원본XML 매핑에 있으나 실제 워크북에 없는 시트: {missing[:8]}')
    merged_cnt = sum(len(wb[sn].merged_cells.ranges) for sn in visible_targets if sn in wb.sheetnames)
    if merged_cnt:
        issues.append(f'병합셀 {merged_cnt}개가 존재합니다. 일부 수동 편집에서 주의가 필요합니다.')

    return {
        'ok': not fatal,
        'issues': issues,
        'repairable': repairable,
        'fatal': fatal,
        'sheet_count': len(visible_targets),
    }

def repair_excel_bytes_for_upload(xlsx_bytes):
    """보수적 복구만 수행. 핵심 구조를 추정 복원하지는 않음."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=False)
    actions = []
    for req in ('_원본XML', '_원본DSD_바이너리'):
        if req in wb.sheetnames and wb[req].sheet_state != 'hidden':
            wb[req].sheet_state = 'hidden'
            actions.append(f'{req} 시트를 hidden 처리')
    # 사용안내 시트가 첫 장이 아니면 앞으로 이동
    if '📋사용안내' in wb.sheetnames and wb.sheetnames[0] != '📋사용안내':
        ws = wb['📋사용안내']
        wb._sheets.remove(ws)
        wb._sheets.insert(0, ws)
        actions.append('📋사용안내 시트를 맨 앞으로 이동')
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue(), actions

def is_note_ref(val):
    parts = str(val).strip().split(',')
    return (len(parts) >= 2 and all(p.strip().isdigit() and 1 <= len(p.strip()) <= 2 for p in parts))

def _is_edit_cell(cell):
    f = cell.fill
    if f and f.fill_type == 'solid':
        fg = f.fgColor
        if fg and fg.type == 'rgb' and isinstance(fg.rgb, str):
            return fg.rgb.upper().endswith(EDIT_COLOR.upper())
    return False

def _rollover_sheet_safe(ws, fill_000=True):
    """dart_gui의 열기반 롤오버 아이디어를 보수적으로 이식."""
    for rowi in range(1, ws.max_row + 1):
        amt_cells = []
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(rowi, ci)
            if not _is_edit_cell(cell) or cell.value is None:
                continue
            raw = str(cell.value).strip()
            if is_note_ref(raw):
                continue
            vclean = raw.replace(',', '').replace('(', '').replace(')', '').replace('-', '').replace(' ', '')
            if vclean and vclean.replace('.', '').isdigit() and len(vclean) >= 4:
                amt_cells.append((ci, cell))
        if len(amt_cells) == 0:
            continue
        if len(amt_cells) == 1:
            _cc, c_cell = amt_cells[0]
            p_cell = None
            for ci2 in range(_cc + 1, ws.max_column + 1):
                cand = ws.cell(rowi, ci2)
                if _is_edit_cell(cand) and str(cand.value or '').strip() in ('-', ''):
                    p_cell = cand
                    break
            if p_cell is None:
                continue
            p_cell.value = c_cell.value
            c_cell.value = '000' if fill_000 else None
            continue
        amt_cells.sort(key=lambda x: x[0])
        _, c_cell = amt_cells[-2]
        _, p_cell = amt_cells[-1]
        p_cell.value = c_cell.value
        c_cell.value = '000' if fill_000 else None

def _apply_rollover_smart_bytes(xlsx_bytes):
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    FIN_KEYWORDS = ('재무상태표', '손익계산서', '포괄손익', '자본변동표', '현금흐름표')
    for sname in wb.sheetnames:
        if sname in ('📋사용안내','_원본XML','_원본DSD_바이너리','🤖AI검증결과'):
            continue
        if '주석' in sname:
            continue
        if any(kw in sname for kw in FIN_KEYWORDS) or sname.startswith(FIN_PREFIXES):
            _rollover_sheet_safe(wb[sname], fill_000=True)
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def _extract_note_num_from_name_or_title(name, title=''):
    text = f'{name} {title}'.strip()
    pats = [r'주석[_\s]*(\d{1,2})', r'(^|[^\d])(\d{1,2})(?:[^\d]|$)']
    for pat in pats:
        m = re.search(pat, text)
        if m:
            val = m.group(1) if m.lastindex == 1 else m.group(2)
            try:
                n = int(val)
                if 1 <= n <= 99:
                    return n
            except:
                pass
    return None

def regroup_note_sheets_bytes(xlsx_bytes, notes_per_sheet=5):
    """앱1 엔진이 만든 개별 주석 시트를 범위별 그룹 시트로 재구성."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    if '_원본XML' not in wb.sheetnames:
        return xlsx_bytes, {'applied': False, 'reason': '_원본XML 없음'}
    dsd_bytes = _read_original_dsd_bytes_from_wb(wb)
    if not dsd_bytes:
        return xlsx_bytes, {'applied': False, 'reason': '원본 DSD 숨김시트 없음'}

    with zipfile.ZipFile(io.BytesIO(dsd_bytes)) as zf:
        xml = zf.read('contents.xml').decode('utf-8', errors='replace')
    _, tables = parse_xml(xml)

    note_conf = detect_note_anchor_conflicts(tables)
    table_to_note = {}
    for a in note_conf['anchors']:
        table_to_note[a['table_idx']] = a['note_num']
    # 직전 앵커 기준으로 비어 있는 번호도 이어받음
    last_n = None
    for t in sorted(tables, key=lambda x: x['idx']):
        if t['idx'] in table_to_note:
            last_n = table_to_note[t['idx']]
        elif not t.get('fin_label') and last_n:
            table_to_note[t['idx']] = last_n

    ws_map = wb['_원본XML']
    note_rows = []
    fin_rows = []
    for row_idx in range(5, ws_map.max_row + 1):
        sn = ws_map.cell(row_idx, 1).value
        t_idx = ws_map.cell(row_idx, 2).value
        typ = ws_map.cell(row_idx, 6).value
        if not sn or sn == '-' or typ != 'TABLE':
            continue
        item = {'row_idx': row_idx, 'sheet': str(sn), 'table_idx': int(t_idx)}
        if str(sn).startswith('📝'):
            item['note_num'] = table_to_note.get(int(t_idx))
            note_rows.append(item)
        else:
            fin_rows.append(item)

    if not note_rows:
        return xlsx_bytes, {'applied': False, 'reason': '주석 시트 없음', 'note_conflicts': note_conf}

    # 그룹핑: note_num 기준, 번호 없으면 뒤로
    ordered = sorted(note_rows, key=lambda x: (999 if x.get('note_num') is None else x['note_num'], x['table_idx']))
    chunks, chunk, current_notes = [], [], []
    for item in ordered:
        n = item.get('note_num')
        if n is not None and n not in current_notes and len(current_notes) >= max(1, notes_per_sheet):
            chunks.append(chunk); chunk = []; current_notes = []
        chunk.append(item)
        if n is not None and n not in current_notes:
            current_notes.append(n)
    if chunk:
        chunks.append(chunk)

    old_note_sheet_names = [x['sheet'] for x in note_rows if x['sheet'] in wb.sheetnames]
    insert_pos = 1 + len([sn for sn in wb.sheetnames if sn.startswith(FIN_PREFIXES)])
    created_names = []

    for ci, items in enumerate(chunks, start=1):
        nums = [x['note_num'] for x in items if x.get('note_num') is not None]
        if nums:
            fn, ln = min(nums), max(nums)
            sname = f'📝주석_{fn}' if fn == ln else f'📝주석_{fn}_{ln}'
        else:
            sname = f'📝주석_기타_{ci:02d}'
        sname = sname[:31]
        base_name = sname
        suffix = 1
        while sname in wb.sheetnames or sname in created_names:
            suffix += 1
            sname = f'{base_name[:28]}_{suffix}'
        ws_new = wb.create_sheet(title=sname, index=min(insert_pos + ci - 1, len(wb.sheetnames)))
        created_names.append(sname)

        cur_row = 1
        for item in items:
            if item['sheet'] not in wb.sheetnames:
                continue
            src = wb[item['sheet']]
            max_col = src.max_column
            for c in range(1, max_col + 1):
                col_letter = get_column_letter(c)
                ws_new.column_dimensions[col_letter].width = src.column_dimensions[col_letter].width
            for r in range(1, src.max_row + 1):
                if src.row_dimensions[r].height:
                    ws_new.row_dimensions[cur_row + r - 1].height = src.row_dimensions[r].height
                for c in range(1, src.max_column + 1):
                    sc = src.cell(r, c)
                    dc = ws_new.cell(cur_row + r - 1, c, sc.value)
                    _copy_cell_style(sc, dc)
            # 매핑 시트의 시트명 갱신
            ws_map.cell(item['row_idx'], 1).value = sname
            cur_row += src.max_row + 2

    for sn in old_note_sheet_names:
        if sn in wb.sheetnames:
            del wb[sn]

    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue(), {
        'applied': True,
        'grouped_sheet_count': len(created_names),
        'group_names': created_names,
        'note_conflicts': note_conf
    }

def apply_period_change_only_bytes(xlsx_bytes, cur_period=None, cur_year=None, year_offset=1,
                                   start_m=None, start_d=None, end_m=None, end_d=None):
    """헤더의 기수/연도/기간 텍스트만 보수적으로 치환."""
    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
    if cur_year is None:
        cur_year = datetime.utcnow().year + max(0, int(year_offset))
    if cur_period is None:
        cur_period = 1

    # dart_gui의 apply_period_change를 보수 이식
    prev_period = int(cur_period) - 1
    prev_year = int(cur_year) - 1
    skip = {'📋사용안내','_원본XML','_원본DSD_바이너리','🤖AI검증결과'}

    old_cur_p = old_prev_p = None
    for sname in wb.sheetnames:
        if not any(sname.startswith(p) for p in FIN_PREFIXES):
            continue
        ws = wb[sname]
        for row in ws.iter_rows(max_row=10, values_only=True):
            for v in row:
                if not v or not isinstance(v, str):
                    continue
                m = re.search(r'제\s*(\d{1,3})\s*\(당\)', v)
                if m and not old_cur_p:
                    old_cur_p = int(m.group(1))
                m = re.search(r'제\s*(\d{1,3})\s*\(전\)', v)
                if m and not old_prev_p:
                    old_prev_p = int(m.group(1))
        if old_cur_p:
            break
    if not old_cur_p or old_cur_p <= 0:
        old_cur_p = cur_period - 1
    if not old_prev_p or old_prev_p <= 0:
        old_prev_p = cur_period - 2

    years = []
    for sname in wb.sheetnames:
        if not any(sname.startswith(p) for p in FIN_PREFIXES):
            continue
        ws = wb[sname]
        for row in ws.iter_rows(max_row=15, values_only=True):
            for v in row:
                if v and isinstance(v, str):
                    years += [int(x) for x in re.findall(r'(20\d{2})년', v)]
        if len(set(years)) >= 2:
            break
    years = sorted(set(years))
    if len(years) >= 2:
        old_prev_y, old_cur_y = years[0], years[1]
    elif len(years) == 1:
        old_cur_y = years[0]; old_prev_y = old_cur_y - 1
    else:
        old_cur_y = cur_year - 1; old_prev_y = cur_year - 2

    def rep(t):
        if not t or not isinstance(t, str):
            return t
        t = re.sub(rf'제\s*{old_cur_p}\s*\(당\)', f'제 {cur_period}(당)', t)
        t = re.sub(rf'제\s*{old_prev_p}\s*\(전\)', f'제 {prev_period}(전)', t)
        t = re.sub(rf'제\s*{old_cur_p}\s*기\b', f'제 {cur_period}기', t)
        t = re.sub(rf'제\s*{old_prev_p}\s*기\b', f'제 {prev_period}기', t)
        t = t.replace(f'{old_cur_y}년', f'{cur_year}년')
        t = t.replace(f'{old_prev_y}년', f'{prev_year}년')
        if all(v is not None for v in (start_m, start_d, end_m, end_d)):
            old_dates = re.findall(r'(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일', t)
            if len(old_dates) >= 1:
                t = re.sub(r'(20\d{2})년\s*\d{1,2}월\s*\d{1,2}일', f'{cur_year}년 {int(start_m)}월 {int(start_d)}일', t, count=1)
            if len(old_dates) >= 2:
                t = re.sub(r'(20\d{2})년\s*\d{1,2}월\s*\d{1,2}일', f'{cur_year}년 {int(end_m)}월 {int(end_d)}일', t, count=1)
        return t

    for sname in wb.sheetnames:
        if sname in skip:
            continue
        ws = wb[sname]
        for rowi in range(1, ws.max_row + 1):
            for ci in range(1, ws.max_column + 1):
                cell = ws.cell(rowi, ci)
                if cell.value and isinstance(cell.value, str):
                    nv = rep(cell.value)
                    if nv != cell.value:
                        cell.value = nv
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

def build_diff_report_xlsx(diff_result):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Diff요약'
    ws.append(['구분', '테이블', '행', '열', '수정전', '수정후'])
    for it in diff_result.get('changed', []):
        ws.append(['변경', it['table'], it['row'], it['col'], it['before'], it['after']])
    for it in diff_result.get('added', []):
        ws.append(['추가', it['table'], it['row'], it['col'], '', it['after']])
    for it in diff_result.get('removed', []):
        ws.append(['삭제', it['table'], it['row'], it['col'], it['before'], ''])
    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = 24 if c >= 5 else 14
    buf = io.BytesIO(); wb.save(buf); return buf.getvalue()

# ── 14. Flask 앱 & HTML 템플릿 ───────────────────────────────────────────────
app = Flask(__name__)

HTML = """<!DOCTYPE html>
<html lang=\"ko\">
<head>
<meta charset=\"UTF-8\">
<meta name=\"viewport\" content=\"width=device-width, initial-scale=1.0\">
<title>EasyDSD v0.13</title>
<style>
*{box-sizing:border-box} body{margin:0;font-family:'Malgun Gothic',sans-serif;background:#e9edf2;color:#223}
.hd{background:linear-gradient(135deg,#0f2845,#244b73);color:#fff;padding:18px 24px}
.hd h1{margin:0;font-size:22px}.hd p{margin:6px 0 0;font-size:12px;opacity:.88}
.api{background:#fff;padding:12px 24px;border-bottom:1px solid #d7dfe8;display:flex;gap:10px;align-items:center;flex-wrap:wrap}
.api label{font-size:12px;font-weight:bold;color:#556}.api input,.api select{padding:7px 10px;border:1px solid #c5d0db;border-radius:8px}
.api input{min-width:240px}.badge{display:inline-block;padding:5px 10px;border-radius:999px;background:#eef2f6;color:#667;font-size:12px;font-weight:bold}
.badge.ok{background:#d4edda;color:#155724}.wrap{max-width:1100px;margin:18px auto;padding:0 14px 30px}
.tabs{display:flex;gap:6px;flex-wrap:wrap}.tb{border:none;border-radius:12px 12px 0 0;background:#c8d3de;color:#455;padding:11px 14px;font-weight:bold;cursor:pointer}.tb.on{background:#fff;color:#163a5e}
.pane{display:none;background:#fff;border-radius:0 14px 14px 14px;padding:22px;box-shadow:0 3px 16px rgba(0,0,0,.07)}.pane.on{display:block}
.notice{border-radius:10px;padding:12px 15px;margin-bottom:16px;border:2px solid #ffc107;background:linear-gradient(135deg,#fff3cd,#ffe39c)} .notice.red{border-color:#dc3545;background:linear-gradient(135deg,#f8d7da,#f3bec5)}
.notice strong{display:block;color:#6c4b00;font-size:14px}.notice.red strong{color:#842029}.notice p{margin:4px 0 0;font-size:12px}
.dz{border:2px dashed #96a9bf;border-radius:12px;padding:26px;text-align:center;background:#f7fafe;cursor:pointer;margin:10px 0;transition:.15s ease}.dz.ok{border-color:#28a745;background:#f2fff5}.dz.drag{border-color:#0d6efd;background:#eef6ff;transform:translateY(-1px)}.dz .ico{font-size:26px}
.grid2{display:grid;grid-template-columns:1fr 1fr;gap:14px}.grid3{display:grid;grid-template-columns:repeat(3,1fr);gap:12px}
.opt{background:#f3f7fb;border-radius:10px;padding:14px;margin:12px 0}.opt label{display:flex;gap:8px;align-items:flex-start;margin:8px 0;font-size:13px}
.opt small{color:#667}.opt input[type=number], .opt select{padding:6px 8px;border:1px solid #c5d0db;border-radius:8px}
.btn{width:100%;padding:14px;border:none;border-radius:12px;color:#fff;font-size:15px;font-weight:bold;cursor:pointer;margin-top:12px}
.btn:disabled{opacity:.55;cursor:not-allowed}.b1{background:linear-gradient(135deg,#275d8f,#2ba7a6)}.b2{background:linear-gradient(135deg,#7b46d3,#9157d9)}.b3{background:linear-gradient(135deg,#1f8d4a,#2bbb75)}
.b4{background:linear-gradient(135deg,#db7a12,#ef9a1a)}.b5{background:linear-gradient(135deg,#7d3fc7,#b05ee5)}.b6{background:linear-gradient(135deg,#34495e,#4f6b85)}
.spin{display:none;text-align:center;padding:16px}.spin.on{display:block}.res{display:none;margin-top:14px}.box{background:#f8fafc;border:1px solid #d7e0e9;border-radius:10px;padding:14px;white-space:pre-wrap;overflow:auto;max-height:520px;font-size:12px;line-height:1.7}
.dl{display:inline-block;margin-top:10px;padding:11px 18px;border-radius:10px;background:#0d6efd;color:#fff;text-decoration:none;font-weight:bold}
.mini{font-size:12px;color:#667}.kv{display:grid;grid-template-columns:180px 1fr;gap:10px;font-size:12px;margin:6px 0}.kv b{color:#28496a}
.tbl{width:100%;border-collapse:collapse;font-size:12px}.tbl th{background:#1f4e79;color:#fff;padding:8px}.tbl td{padding:7px 8px;border-bottom:1px solid #e5ebf2;vertical-align:top}
.sec h3{margin:0 0 10px;color:#163a5e}.pill{display:inline-block;padding:3px 9px;border-radius:99px;font-size:11px;font-weight:bold}.ok{background:#d4edda;color:#155724}.wn{background:#fff3cd;color:#856404}.er{background:#f8d7da;color:#842029}
.devgrid{display:grid;grid-template-columns:1.1fr .9fr;gap:18px}.card{background:#f6f9fc;border:1px solid #dbe4ec;border-radius:12px;padding:16px}
ul.clean{margin:8px 0 0 18px;padding:0} ul.clean li{margin:6px 0}
@media (max-width:900px){.grid2,.grid3,.devgrid{grid-template-columns:1fr}.api input{min-width:180px}}
</style>
</head>
<body>
<div class=\"hd\">
  <h1>⚡ EasyDSD v0.13</h1>
  <p>app(1) 핵심 DSD↔Excel 엔진 유지 + 롤오버 + 주석정리 + 기수/연도 밀기 + AI검증 + 전기금액검증 + Diff</p>
</div>

<div class=\"api\">
  <label>AI 제공자</label>
  <select id=\"apiProvider\">
    <option value=\"gemini\">Gemini</option>
    <option value=\"anthropic\">Anthropic (Claude)</option>
  </select>
  <label>API Key</label>
  <input id=\"apiKey\" type=\"password\" placeholder=\"AI 검증 탭에서만 사용\">
  <label>모델</label>
  <select id=\"apiModel\">
    <option value=\"gemini-1.5-flash\">Gemini 1.5 Flash</option>
    <option value=\"gemini-1.5-pro\">Gemini 1.5 Pro</option>
    <option value=\"gemini-2.0-flash\">Gemini 2.0 Flash</option>
    <option value=\"claude-3-5-sonnet-latest\">Claude 3.5 Sonnet</option>
    <option value=\"claude-3-7-sonnet-latest\">Claude 3.7 Sonnet</option>
  </select>
  <span id=\"apiBadge\" class=\"badge\">미설정</span>
</div>

<div class=\"wrap\">
  <div class=\"tabs\">
    <button class=\"tb on\" onclick=\"sw('p1',this)\">① DSD→Excel</button>
    <button class=\"tb\" onclick=\"sw('p2',this)\">② AI 검증</button>
    <button class=\"tb\" onclick=\"sw('p3',this)\">③ Excel→DSD</button>
    <button class=\"tb\" onclick=\"sw('p4',this)\">④ 전기금액 검증</button>
    <button class=\"tb\" onclick=\"sw('p5',this)\">⑤ DSD 비교분석</button>
    <button class=\"tb\" onclick=\"sw('p6',this)\">⑥ 개발자 정보</button>
  </div>

  <div id=\"p1\" class=\"pane on\">
    <div class=\"notice\">
      <strong>⚠ 행 추가 시 [복사된 셀 삽입] 필수</strong>
      <p>핵심 엔진이 행 추가/삭제를 반영하려면 기존 행을 복사한 뒤 <b>복사된 셀 삽입</b>으로 작업해야 합니다.</p>
    </div>
    <div class=\"dz\" id=\"dz1\" onclick=\"pick('f1')\"><div class=\"ico\">📁</div><div id=\"dt1\">DSD 파일을 선택하세요</div></div>
    <input id=\"f1\" type=\"file\" accept=\".dsd,.zip\" style=\"display:none\" onchange=\"loadFile('1')\">

    <div class=\"opt\">
      <label><input type=\"checkbox\" id=\"optRollover\"> <div><b>전기로 밀기(롤오버)</b><br><small>당기 숫자를 전기 칸으로 밀고 당기 칸을 000으로 채웁니다.</small></div></label>
      <label><input type=\"checkbox\" id=\"optNotes\" checked> <div><b>분리된 시트 정리</b><br><small>주석 시트를 번호 범위 기준으로 묶어 정리합니다. _원본XML 매핑도 함께 갱신합니다.</small></div></label>
      <label><input type=\"checkbox\" id=\"optPeriod\"> <div><b>기수/연도 밀기</b><br><small>헤더의 제 n기 / 연도 문구를 보수적으로 갱신합니다.</small></div></label>
      <div class=\"grid3\">
        <div><small>연도 오프셋</small><br><input type=\"number\" id=\"yearOffset\" value=\"1\" min=\"-5\" max=\"5\"></div>
        <div><small>주석 묶음 크기</small><br><select id=\"noteChunk\"><option value=\"3\">3개 주석씩</option><option value=\"5\" selected>5개 주석씩</option><option value=\"7\">7개 주석씩</option></select></div>
        <div><small>현재 기수(선택)</small><br><input type=\"number\" id=\"curPeriod\" value=\"1\" min=\"1\" max=\"999\"></div>
      </div>
      <div class=\"grid3\" style=\"margin-top:8px\">
        <div><small>현재 연도(선택)</small><br><input type=\"number\" id=\"curYear\" value=\"2026\" min=\"2000\" max=\"2099\"></div>
        <div><small>시작월(선택)</small><br><input type=\"number\" id=\"startM\" min=\"1\" max=\"12\"></div>
        <div><small>시작일(선택)</small><br><input type=\"number\" id=\"startD\" min=\"1\" max=\"31\"></div>
      </div>
      <div class=\"grid3\" style=\"margin-top:8px\">
        <div><small>종료월(선택)</small><br><input type=\"number\" id=\"endM\" min=\"1\" max=\"12\"></div>
        <div><small>종료일(선택)</small><br><input type=\"number\" id=\"endD\" min=\"1\" max=\"31\"></div>
        <div><small>변환 전 점검</small><br><button type=\"button\" id=\"btnPrecheck\" class=\"btn b6\" style=\"margin-top:0;padding:10px\" onclick=\"precheckDsd()\">사전 점검</button></div>
      </div>
    </div>
    <button class=\"btn b1\" id=\"btn1\" onclick=\"runDsd2Excel()\" disabled>📊 Excel 파일 생성</button>
    <div id=\"sp1\" class=\"spin\">변환 중...</div>
    <div id=\"r1\" class=\"res\"><div id=\"r1box\" class=\"box\"></div></div>
  </div>

  <div id=\"p2\" class=\"pane\">
    <div class=\"notice red\">
      <strong>🤖 AI 검증 탭은 API Key 입력 사용자만 실행</strong>
      <p>여기서 대차평균, 합계/총계 행 존재 여부, 자릿수 편차를 Python으로 계산한 뒤 AI 해설을 붙입니다. 합계/총계 관련 검증도 이 탭에서만 수행합니다.</p>
    </div>
    <div class=\"dz\" id=\"dz2\" onclick=\"pick('f2')\"><div class=\"ico\">🤖</div><div id=\"dt2\">검증할 Excel(.xlsx) 파일</div></div>
    <input id=\"f2\" type=\"file\" accept=\".xlsx\" style=\"display:none\" onchange=\"loadFile('2')\">
    <button class=\"btn b2\" id=\"btn2\" onclick=\"runAiValidate()\" disabled>🤖 AI 검증 실행</button>
    <div id=\"sp2\" class=\"spin\">AI 검증 중...</div>
    <div id=\"r2\" class=\"res\"><div id=\"r2box\" class=\"box\"></div></div>
  </div>

  <div id=\"p3\" class=\"pane\">
    <div class=\"notice\">
      <strong>⚠ 핵심 재조립 전 사전 점검 권장</strong>
      <p>업로드 가능 여부를 먼저 점검하고, 복구 가능한 항목은 별도 복구본을 내려받아 다시 시도할 수 있습니다.</p>
    </div>
    <div class=\"dz\" id=\"dz3\" onclick=\"pick('f3')\"><div class=\"ico\">📄</div><div id=\"dt3\">수정된 Excel(.xlsx) 파일</div></div>
    <input id=\"f3\" type=\"file\" accept=\".xlsx\" style=\"display:none\" onchange=\"loadFile('3')\">
    <div class=\"grid2\">
      <button class=\"btn b6\" id=\"btn3a\" onclick=\"runExcelPrecheck()\" disabled>🩺 업로드 가능 여부 점검</button>
      <button class=\"btn b6\" id=\"btn3b\" onclick=\"runExcelRepair()\" disabled>🛠 복구본 생성</button>
    </div>
    <button class=\"btn b3\" id=\"btn3\" onclick=\"runExcel2Dsd()\" disabled>🔧 DSD 파일 재조립</button>
    <div id=\"sp3\" class=\"spin\">재조립 중...</div>
    <div id=\"r3\" class=\"res\"><div id=\"r3box\" class=\"box\"></div></div>
  </div>

  <div id=\"p4\" class=\"pane\">
    <div class=\"notice red\">
      <strong>📋 전기금액 검증</strong>
      <p>작년 DSD의 당기 금액과 올해 DSD의 전기 금액을 항목별로 대조합니다.</p>
    </div>
    <div class=\"grid2\">
      <div>
        <div class=\"dz\" id=\"dz4p\" onclick=\"pick('f4p')\"><div class=\"ico\">📁</div><div id=\"dt4p\">작년 DSD</div></div>
        <input id=\"f4p\" type=\"file\" accept=\".dsd,.zip\" style=\"display:none\" onchange=\"loadFile('4p')\">
      </div>
      <div>
        <div class=\"dz\" id=\"dz4c\" onclick=\"pick('f4c')\"><div class=\"ico\">📁</div><div id=\"dt4c\">올해 DSD</div></div>
        <input id=\"f4c\" type=\"file\" accept=\".dsd,.zip\" style=\"display:none\" onchange=\"loadFile('4c')\">
      </div>
    </div>
    <button class=\"btn b4\" id=\"btn4\" onclick=\"runPrior()\" disabled>🔍 전기금액 검증</button>
    <div id=\"sp4\" class=\"spin\">비교 중...</div>
    <div id=\"r4\" class=\"res\"><div id=\"r4box\" class=\"box\"></div></div>
  </div>

  <div id=\"p5\" class=\"pane\">
    <div class=\"notice red\">
      <strong>🔎 DSD 비교분석</strong>
      <p>변경된 셀만 추출해 보여주고, 원하면 Diff 리포트를 Excel로도 내려받을 수 있습니다.</p>
    </div>
    <div class=\"grid2\">
      <div>
        <div class=\"dz\" id=\"dz5b\" onclick=\"pick('f5b')\"><div class=\"ico\">📁</div><div id=\"dt5b\">수정 전 DSD</div></div>
        <input id=\"f5b\" type=\"file\" accept=\".dsd,.zip\" style=\"display:none\" onchange=\"loadFile('5b')\">
      </div>
      <div>
        <div class=\"dz\" id=\"dz5a\" onclick=\"pick('f5a')\"><div class=\"ico\">📁</div><div id=\"dt5a\">수정 후 DSD</div></div>
        <input id=\"f5a\" type=\"file\" accept=\".dsd,.zip\" style=\"display:none\" onchange=\"loadFile('5a')\">
      </div>
    </div>
    <div class=\"grid2\">
      <button class=\"btn b5\" id=\"btn5\" onclick=\"runDiff()\" disabled>🔎 JSON 비교 보기</button>
      <button class=\"btn b6\" id=\"btn5x\" onclick=\"downloadDiffXlsx()\" disabled>📥 Diff 리포트 다운로드</button>
    </div>
    <div id=\"sp5\" class=\"spin\">비교 중...</div>
    <div id=\"r5\" class=\"res\"><div id=\"r5box\" class=\"box\"></div></div>
  </div>

  <div id=\"p6\" class=\"pane\">
    <div class=\"devgrid\">
      <div class=\"card\">
        <h2 style=\"margin-top:0\">개발자 정보</h2>
        <div class=\"kv\"><b>프로젝트</b><div>EasyDSD v0.13</div></div>
        <div class=\"kv\"><b>베이스 엔진</b><div>app(1)의 DSD↔Excel 핵심 엔진 유지</div></div>
        <div class=\"kv\"><b>추가 레이어</b><div>롤오버 · 주석시트 정리 · 기수/연도 변경 · AI검증 · 사전점검 · 복구본 생성 · Diff 리포트</div></div>
        <div class=\"kv\"><b>지원 파일</b><div>.dsd / .xlsx</div></div>
        <div class=\"kv\"><b>연락</b><div><a href=\"mailto:eeffco11@naver.com\">eeffco11@naver.com</a></div></div>
      </div>
      <div class=\"card\">
        <h3 style=\"margin-top:0\">안전 원칙</h3>
        <ul class=\"clean\">
          <li>핵심 DSD↔Excel 엔진은 직접 수정하지 않음</li>
          <li>부가기능은 후처리/검증/리포트 레이어로만 추가</li>
          <li>스타일을 대량 재작성하는 기능은 제외</li>
          <li>행 추가는 반드시 복사된 셀 삽입 방식 권장</li>
        </ul>
      </div>
    </div>
    <div class=\"card\" style=\"margin-top:16px\">
      <h3 style=\"margin-top:0\">이번 버전 주요 기능</h3>
      <ul class=\"clean\">
        <li>DSD→Excel: 롤오버 / 주석 시트 정리 / 기수·연도 밀기</li>
        <li>AI 검증: API Key 입력 사용자만 실행, 합계·총계 관련 검증 포함</li>
        <li>Excel→DSD: 사전 점검 및 제한적 복구본 생성</li>
        <li>전기금액 검증 및 DSD Diff 리포트 다운로드</li>
        <li>주석 번호 충돌 탐지 및 사용자 친화 오류 메시지</li>
      </ul>
    </div>
  </div>
</div>

<script>
(function(){
  'use strict';
  const $ = (id) => document.getElementById(id);
  const F = {};
  const DZ = {
    '1': {inp:'f1', txt:'dt1', dz:'dz1', btns:['btn1']},
    '2': {inp:'f2', txt:'dt2', dz:'dz2', btns:['btn2']},
    '3': {inp:'f3', txt:'dt3', dz:'dz3', btns:['btn3','btn3a','btn3b']},
    '4p':{inp:'f4p',txt:'dt4p',dz:'dz4p',btns:[]},
    '4c':{inp:'f4c',txt:'dt4c',dz:'dz4c',btns:[]},
    '5b':{inp:'f5b',txt:'dt5b',dz:'dz5b',btns:[]},
    '5a':{inp:'f5a',txt:'dt5a',dz:'dz5a',btns:[]}
  };

  function esc(s){ return String(s ?? '').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }
  function byIdOrNull(id){ return document.getElementById(id); }
  function setText(id, text){ const el=byIdOrNull(id); if(el) el.textContent=text; }
  function setHtml(id, html){ const el=byIdOrNull(id); if(el) el.innerHTML=html; }
  function show(id, on=true){ const el=byIdOrNull(id); if(el) el.style.display = on ? 'block' : 'none'; }
  function toggleDisabled(id, on){ const el=byIdOrNull(id); if(el) el.disabled = !!on; }
  function getResultBox(n){ return {wrap: byIdOrNull('r'+n), box: byIdOrNull('r'+n+'box') || byIdOrNull('r'+n)}; }

  function refreshPairButtons(){
    toggleDisabled('btn4', !(F['4p'] && F['4c']));
    const diffReady = !!(F['5b'] && F['5a']);
    toggleDisabled('btn5', !diffReady);
    toggleDisabled('btn5x', !diffReady);
  }

  function loadFile(key, files){
    const cfg = DZ[key];
    if(!cfg || !files || !files[0]) return;
    F[key] = files[0];
    setText(cfg.txt, '✅ ' + F[key].name);
    const dz = byIdOrNull(cfg.dz);
    if(dz) dz.classList.add('ok');
    (cfg.btns || []).forEach(id => toggleDisabled(id, false));
    refreshPairButtons();
  }

  function bindDropzone(key){
    const cfg = DZ[key];
    const dz = byIdOrNull(cfg.dz), inp = byIdOrNull(cfg.inp);
    if(!dz || !inp) return;
    dz.addEventListener('click', (e)=>{
      if(e.target && e.target.closest('input,button,a,label,select,option,textarea')) return;
      inp.click();
    });
    inp.addEventListener('change', ()=>loadFile(key, inp.files));
    ['dragenter','dragover'].forEach(ev => dz.addEventListener(ev, (e)=>{ e.preventDefault(); e.stopPropagation(); dz.classList.add('drag'); }));
    ['dragleave','dragend'].forEach(ev => dz.addEventListener(ev, (e)=>{ e.preventDefault(); e.stopPropagation(); dz.classList.remove('drag'); }));
    dz.addEventListener('drop', (e)=>{
      e.preventDefault(); e.stopPropagation(); dz.classList.remove('drag');
      const files = e.dataTransfer && e.dataTransfer.files;
      if(!files || !files.length) return;
      try {
        const dt = new DataTransfer();
        dt.items.add(files[0]);
        inp.files = dt.files;
      } catch(_e) {}
      loadFile(key, files);
    });
  }

  function setBusy(n,on,msg){
    const sp = byIdOrNull('sp'+n);
    if(sp){
      sp.classList[on?'add':'remove']('on');
      if(msg) sp.textContent = msg;
    }
    toggleDisabled('btn'+n, on);
  }

  function renderPrecheck(info){
    let h='<h3>사전 점검 결과</h3>';
    h+='<div class="kv"><b>전체 TABLE</b><div>'+esc(info.table_count)+'</div></div>';
    h+='<div class="kv"><b>재무제표 TABLE</b><div>'+esc(info.financial_table_count)+'</div></div>';
    h+='<div class="kv"><b>주석 TABLE</b><div>'+esc(info.note_table_count)+'</div></div>';
    if(info.warnings&&info.warnings.length) h+='<div><span class="pill wn">주의</span> '+esc(info.warnings.join(' / '))+'</div>';
    const nc=info.note_conflicts||{};
    h+='<div class="kv"><b>주석 중복</b><div>'+esc((nc.duplicates||[]).length)+'건</div></div>';
    h+='<div class="kv"><b>주석 공백 번호</b><div>'+esc(((nc.missing_numbers||[]).join(', '))||'없음')+'</div></div>';
    return h;
  }
  function renderAi(data){
    let h='<h3>Python 검증 요약</h3>';
    const s=(data.python_validation||{}).summary||{};
    h+='<div class="kv"><b>시트 수</b><div>'+esc(s.sheet_count)+'</div></div>';
    h+='<div class="kv"><b>숫자 셀 수</b><div>'+esc(s.numeric_cell_count)+'</div></div>';
    h+='<div class="kv"><b>경고</b><div>'+esc(s.warning_count)+'</div></div>';
    h+='<div class="kv"><b>오류</b><div>'+esc(s.error_count)+'</div></div>';
    h+='<hr><h3>AI 의견</h3><div>'+esc(data.ai_validation||'').replace(/\n/g,'<br>')+'</div>';
    return h;
  }
  function renderPrior(data){
    let h='<h3>전기금액 검증 결과</h3>';
    h+='<div class="kv"><b>일치</b><div>'+esc(data.summary?.ok)+'</div></div>';
    h+='<div class="kv"><b>불일치</b><div>'+esc(data.summary?.mismatch)+'</div></div>';
    h+='<div class="kv"><b>누락</b><div>'+esc(data.summary?.missing)+'</div></div>';
    h+='<table class="tbl"><tr><th>재무제표</th><th>샘플</th></tr>';
    (data.tables||[]).slice(0,20).forEach(t=>{ const row=(t.rows||[]).find(x=>x.status!=='✅일치') || (t.rows||[])[0]; h+='<tr><td>'+esc(t.label)+'</td><td>'+esc(row ? (row.item+' / '+row.status) : '')+'</td></tr>'; });
    h+='</table>'; return h;
  }
  function renderDiff(data){
    let h='<h3>DSD Diff 요약</h3>';
    h+='<div class="kv"><b>변경</b><div>'+esc(data.summary?.changed_count)+'</div></div>';
    h+='<div class="kv"><b>추가</b><div>'+esc(data.summary?.added_count)+'</div></div>';
    h+='<div class="kv"><b>삭제</b><div>'+esc(data.summary?.removed_count)+'</div></div>';
    h+='<hr><pre>'+esc(JSON.stringify(data,null,2))+'</pre>';
    return h;
  }
  function lsGet(k,d){ try{ const v=window.localStorage ? localStorage.getItem(k) : null; return v===null ? d : v; }catch(_){ return d; } }
  function lsSet(k,v){ try{ if(window.localStorage) localStorage.setItem(k,v); }catch(_){ } }
  function saveApiPrefs(){
    const apiKey = byIdOrNull('apiKey'), apiProvider = byIdOrNull('apiProvider'), apiModel = byIdOrNull('apiModel'), apiBadge = byIdOrNull('apiBadge');
    if(apiKey) lsSet('easydsd_api_key', apiKey.value);
    if(apiProvider) lsSet('easydsd_provider', apiProvider.value);
    if(apiModel) lsSet('easydsd_model', apiModel.value);
    if(apiBadge && apiKey){
      const on = apiKey.value.trim().length > 8;
      apiBadge.textContent = on ? '설정됨 ✓' : '미설정';
      apiBadge.className = 'badge' + (on ? ' ok' : '');
    }
  }

  window.sw = function(id, btn){
    document.querySelectorAll('.pane').forEach(x=>x.classList.remove('on'));
    document.querySelectorAll('.tb').forEach(x=>x.classList.remove('on'));
    const pane = byIdOrNull(id); if(pane) pane.classList.add('on');
    if(btn) btn.classList.add('on');
  };
  window.pick = function(id){ const inp = byIdOrNull(id); if(inp) inp.click(); };
  window.loadFile = function(key){ const cfg=DZ[key]; const inp=cfg?byIdOrNull(cfg.inp):null; if(inp) loadFile(key, inp.files); };

  window.precheckDsd = async function(){
    if(!F['1']) return alert('DSD 파일을 먼저 선택하세요.');
    const fd=new FormData(); fd.append('dsd',F['1']);
    const r=await fetch('/api/precheck_dsd',{method:'POST',body:fd});
    const data=await r.json(); const out=getResultBox(1); show('r1', true); setHtml(out.box.id, r.ok?renderPrecheck(data):('❌ '+esc(data.error||'사전 점검 실패')));
  };

  window.runDsd2Excel = async function(){
    if(!F['1']) return alert('DSD 파일을 먼저 선택하세요.');
    setBusy(1,true,'변환 중...');
    try{
      const fd=new FormData();
      fd.append('dsd', F['1']);
      fd.append('rollover', byIdOrNull('optRollover')?.checked ? '1' : '0');
      fd.append('organize_notes', byIdOrNull('optNotes')?.checked ? '1' : '0');
      fd.append('period_change', byIdOrNull('optPeriod')?.checked ? '1' : '0');
      fd.append('year_offset', byIdOrNull('yearOffset')?.value || '1');
      fd.append('notes_per_sheet', byIdOrNull('noteChunk')?.value || '5');
      fd.append('cur_period', byIdOrNull('curPeriod')?.value || '');
      fd.append('cur_year', byIdOrNull('curYear')?.value || '');
      fd.append('start_m', byIdOrNull('startM')?.value || '');
      fd.append('start_d', byIdOrNull('startD')?.value || '');
      fd.append('end_m', byIdOrNull('endM')?.value || '');
      fd.append('end_d', byIdOrNull('endD')?.value || '');
      const r=await fetch('/api/dsd2excel',{method:'POST',body:fd});
      if(!r.ok){ const e=await r.json(); throw new Error(e.error||'변환 실패'); }
      const blob=await r.blob(); const info=JSON.parse(r.headers.get('X-Info')||'{}'); const url=URL.createObjectURL(blob);
      let html='<h3>변환 결과 요약</h3>';
      html+='<div class="kv"><b>생성 시트 수</b><div>'+esc(info.sheet_count ?? '-')+'</div></div>';
      html+='<div class="kv"><b>재무 시트 수</b><div>'+esc(info.fin_sheet_count ?? '-')+'</div></div>';
      html+='<div class="kv"><b>주석 시트 수</b><div>'+esc(info.note_sheet_count ?? '-')+'</div></div>';
      html+='<div class="kv"><b>롤오버</b><div>'+(info.rollover?'적용':'미적용')+'</div></div>';
      html+='<div class="kv"><b>주석 정리</b><div>'+(info.organize_notes?'적용':'미적용')+'</div></div>';
      html+='<div class="kv"><b>기수/연도 밀기</b><div>'+(info.period_change?'적용':'미적용')+'</div></div>';
      if(info.note_conflicts) html+='<div class="kv"><b>주석 번호 중복</b><div>'+esc(info.note_conflicts.duplicates || 0)+'건</div></div>';
      html+='<a class="dl" download="작업양식.xlsx" href="'+url+'">📥 Excel 다운로드</a>';
      show('r1', true); setHtml('r1box', html);
    } catch(e){ show('r1', true); setHtml('r1box', '❌ '+esc(e.message)); }
    finally{ setBusy(1,false); }
  };

  window.runAiValidate = async function(){
    if(!F['2']) return alert('Excel 파일을 먼저 선택하세요.');
    const apiKey = byIdOrNull('apiKey'), apiProvider = byIdOrNull('apiProvider'), apiModel = byIdOrNull('apiModel');
    if(!apiKey || !apiKey.value.trim()) return alert('AI 검증은 API Key 입력 사용자만 실행할 수 있습니다.');
    setBusy(2,true,'AI 검증 중...');
    try{
      const fd=new FormData(); fd.append('xlsx',F['2']); fd.append('api_key',apiKey.value.trim()); fd.append('provider',apiProvider?.value||'gemini'); fd.append('model',apiModel?.value||'');
      const r=await fetch('/api/ai_validate',{method:'POST',body:fd}); const data=await r.json();
      show('r2', true); setHtml('r2box', r.ok?renderAi(data):('❌ '+esc(data.error||'검증 실패')));
    } finally { setBusy(2,false); }
  };

  window.runExcelPrecheck = async function(){
    if(!F['3']) return alert('Excel 파일을 먼저 선택하세요.');
    setBusy(3,true,'사전 점검 중...');
    try{
      const fd=new FormData(); fd.append('xlsx',F['3']);
      const r=await fetch('/api/precheck_excel',{method:'POST',body:fd}); const data=await r.json();
      let html=r.ok?'<h3>업로드 사전 점검</h3>':'<h3>점검 실패</h3>';
      if(r.ok){
        html+='<div class="kv"><b>업로드 가능</b><div>'+(data.ok?'예':'아니오')+'</div></div>';
        html+='<div class="kv"><b>편집 시트 수</b><div>'+esc(data.sheet_count)+'</div></div>';
        html+='<div><b>이슈</b><br>'+esc((data.issues||[]).join('\n')||'없음').replace(/\n/g,'<br>')+'</div><br>';
        html+='<div><b>복구 가능 항목</b><br>'+esc((data.repairable||[]).join('\n')||'없음').replace(/\n/g,'<br>')+'</div><br>';
        html+='<div><b>치명 항목</b><br>'+esc((data.fatal||[]).join('\n')||'없음').replace(/\n/g,'<br>')+'</div>';
      } else html+='❌ '+esc(data.error||'점검 실패');
      show('r3', true); setHtml('r3box', html);
    } finally { setBusy(3,false); }
  };

  window.runExcelRepair = async function(){
    if(!F['3']) return alert('Excel 파일을 먼저 선택하세요.');
    setBusy(3,true,'복구본 생성 중...');
    try{
      const fd=new FormData(); fd.append('xlsx',F['3']);
      const r=await fetch('/api/repair_excel',{method:'POST',body:fd});
      if(!r.ok){ const e=await r.json(); throw new Error(e.error||'복구 실패'); }
      const blob=await r.blob(); const info=JSON.parse(r.headers.get('X-Info')||'{}'); const url=URL.createObjectURL(blob);
      show('r3', true); setHtml('r3box', '<h3>복구본 생성 완료</h3><div>'+esc((info.actions||[]).join(' / ')||'적용 가능한 보수적 복구만 수행')+'</div><a class="dl" href="'+url+'" download="복구본.xlsx">📥 복구본 다운로드</a>');
    } catch(e){ show('r3', true); setHtml('r3box', '❌ '+esc(e.message)); }
    finally { setBusy(3,false); }
  };

  window.runExcel2Dsd = async function(){
    if(!F['3']) return alert('Excel 파일을 먼저 선택하세요.');
    setBusy(3,true,'재조립 중...');
    try{
      const fd=new FormData(); fd.append('xlsx',F['3']);
      const r=await fetch('/api/excel2dsd',{method:'POST',body:fd});
      if(!r.ok){ const e=await r.json(); throw new Error(e.error||'재조립 실패'); }
      const blob=await r.blob(); const url=URL.createObjectURL(blob); const info=JSON.parse(r.headers.get('X-Info')||'{}');
      show('r3', true); setHtml('r3box', '<h3>DSD 재조립 완료</h3><div class="kv"><b>상태</b><div>'+esc(info.status||'ok')+'</div></div><a class="dl" href="'+url+'" download="조립결과.dsd">📥 DSD 다운로드</a>');
    } catch(e){ show('r3', true); setHtml('r3box', '❌ '+esc(e.message)); }
    finally { setBusy(3,false); }
  };

  window.runPrior = async function(){
    if(!(F['4p'] && F['4c'])) return alert('두 개의 DSD 파일을 모두 선택하세요.');
    setBusy(4,true,'전기금액 비교 중...');
    try{
      const fd=new FormData(); fd.append('prev_dsd',F['4p']); fd.append('curr_dsd',F['4c']);
      const r=await fetch('/api/verify_prior',{method:'POST',body:fd}); const data=await r.json();
      show('r4', true); setHtml('r4box', r.ok?renderPrior(data):('❌ '+esc(data.error||'검증 실패')));
    } finally { setBusy(4,false); }
  };

  window.runDiff = async function(){
    if(!(F['5b'] && F['5a'])) return alert('두 개의 DSD 파일을 모두 선택하세요.');
    setBusy(5,true,'DSD 비교 중...');
    try{
      const fd=new FormData(); fd.append('before_dsd',F['5b']); fd.append('after_dsd',F['5a']);
      const r=await fetch('/api/diff_dsd',{method:'POST',body:fd}); const data=await r.json();
      show('r5', true); setHtml('r5box', r.ok?renderDiff(data):('❌ '+esc(data.error||'비교 실패')));
    } finally { setBusy(5,false); }
  };

  window.downloadDiffXlsx = async function(){
    if(!(F['5b'] && F['5a'])) return alert('두 개의 DSD 파일을 모두 선택하세요.');
    const fd=new FormData(); fd.append('before_dsd',F['5b']); fd.append('after_dsd',F['5a']);
    const r=await fetch('/api/diff_dsd_xlsx',{method:'POST',body:fd});
    if(!r.ok){ const e=await r.json(); return alert(e.error||'다운로드 실패'); }
    const blob=await r.blob(); const url=URL.createObjectURL(blob); const a=document.createElement('a'); a.href=url; a.download='DSD_Diff_Report.xlsx'; document.body.appendChild(a); a.click(); a.remove();
  };

  document.addEventListener('DOMContentLoaded', function(){
    ['1','2','3','4p','4c','5b','5a'].forEach(bindDropzone);

    document.querySelectorAll('[onclick],[onchange]').forEach(el => {
      el.removeAttribute('onclick');
      el.removeAttribute('onchange');
    });

    document.querySelectorAll('.tabs .tb').forEach((btn, idx) => {
      btn.addEventListener('click', () => window.sw('p' + String(idx + 1), btn));
    });

    const clickMap = {
      'btn1': window.runDsd2Excel,
      'btn2': window.runAiValidate,
      'btn3': window.runExcel2Dsd,
      'btn3a': window.runExcelPrecheck,
      'btn3b': window.runExcelRepair,
      'btn4': window.runPrior,
      'btn5': window.runDiff,
      'btn5x': window.downloadDiffXlsx,
      'btnPrecheck': window.precheckDsd,
    };
    Object.entries(clickMap).forEach(([id, fn]) => {
      const el = byIdOrNull(id);
      if(el && typeof fn === 'function') el.addEventListener('click', fn);
    });
    // ※ pickMap 블록 제거: bindDropzone()가 이미 각 드롭존에 클릭 핸들러를 등록하므로
    //   여기서 또 등록하면 inp.click()이 두 번 호출되어 파일 선택창이 열리지 않음.

    const apiKey = byIdOrNull('apiKey'), apiProvider = byIdOrNull('apiProvider'), apiModel = byIdOrNull('apiModel');
    if(apiKey){ apiKey.value = lsGet('easydsd_api_key', ''); apiKey.addEventListener('input', saveApiPrefs); }
    if(apiProvider){ apiProvider.value = lsGet('easydsd_provider', 'gemini'); apiProvider.addEventListener('change', saveApiPrefs); }
    if(apiModel){ apiModel.value = lsGet('easydsd_model', 'gemini-1.5-flash'); apiModel.addEventListener('change', saveApiPrefs); }
    saveApiPrefs();
    refreshPairButtons();
  });
})();
</script>
</body>
</html>"""

@app.route('/')
def index():
    return Response(HTML, mimetype='text/html; charset=utf-8')

@app.route('/healthz')
def healthz():
    return jsonify({'status':'ok','service':'easydsd','version':'0.13'})

@app.route('/api/precheck_dsd', methods=['POST'])
def api_precheck_dsd():
    try:
        if 'dsd' not in request.files:
            return jsonify({'error':'dsd 파일이 필요합니다.'}), 400
        return jsonify(precheck_dsd_bytes(request.files['dsd'].read()))
    except Exception as e:
        return jsonify({'error': friendly_error_message(e, 'DSD 사전 점검')}), 400

@app.route('/api/precheck_excel', methods=['POST'])
def api_precheck_excel():
    try:
        if 'xlsx' not in request.files:
            return jsonify({'error':'xlsx 파일이 필요합니다.'}), 400
        return jsonify(precheck_excel_bytes(request.files['xlsx'].read()))
    except Exception as e:
        return jsonify({'error': friendly_error_message(e, 'Excel 사전 점검')}), 400

@app.route('/api/repair_excel', methods=['POST'])
def api_repair_excel():
    try:
        if 'xlsx' not in request.files:
            return jsonify({'error':'xlsx 파일이 필요합니다.'}), 400
        fixed, actions = repair_excel_bytes_for_upload(request.files['xlsx'].read())
        resp = send_file(io.BytesIO(fixed),
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name='복구본.xlsx')
        resp.headers['X-Info'] = json.dumps({'actions': actions}, ensure_ascii=True)
        return resp
    except Exception as e:
        return jsonify({'error': friendly_error_message(e, 'Excel 복구본 생성')}), 400

@app.route('/api/dsd2excel', methods=['POST'])
def api_dsd2excel():
    try:
        if 'dsd' not in request.files:
            return jsonify({'error':'dsd 파일이 필요합니다.'}), 400
        dsd_bytes = request.files['dsd'].read()
        do_rollover = request.form.get('rollover','0') == '1'
        organize_notes = request.form.get('organize_notes','0') == '1'
        period_change = request.form.get('period_change','0') == '1'
        year_offset = int(request.form.get('year_offset','1') or '1')
        notes_per_sheet = int(request.form.get('notes_per_sheet','5') or '5')
        cur_period = request.form.get('cur_period','').strip()
        cur_year = request.form.get('cur_year','').strip()
        start_m = request.form.get('start_m','').strip()
        start_d = request.form.get('start_d','').strip()
        end_m = request.form.get('end_m','').strip()
        end_d = request.form.get('end_d','').strip()

        pre = precheck_dsd_bytes(dsd_bytes)
        xlsx = dsd_to_excel_bytes(dsd_bytes, do_period_change=False)

        if do_rollover:
            xlsx = _apply_rollover_smart_bytes(xlsx)
        note_info = {'applied': False}
        if organize_notes:
            xlsx, note_info = regroup_note_sheets_bytes(xlsx, notes_per_sheet=notes_per_sheet)
        if period_change:
            xlsx = apply_period_change_only_bytes(
                xlsx,
                cur_period=int(cur_period) if cur_period else None,
                cur_year=int(cur_year) if cur_year else None,
                year_offset=year_offset,
                start_m=int(start_m) if start_m else None,
                start_d=int(start_d) if start_d else None,
                end_m=int(end_m) if end_m else None,
                end_d=int(end_d) if end_d else None,
            )

        wb = openpyxl.load_workbook(io.BytesIO(xlsx), data_only=False)
        vis = [sn for sn in wb.sheetnames if sn not in ('📋사용안내','_원본XML','_원본DSD_바이너리','🤖AI검증결과')]
        fin_sheets = [sn for sn in vis if sn.startswith(FIN_PREFIXES)]
        note_sheets = [sn for sn in vis if sn.startswith('📝')]
        resp = send_file(io.BytesIO(xlsx),
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='작업양식.xlsx')
        resp.headers['X-Info'] = json.dumps({
            'sheet_count': len(vis),
            'fin_sheet_count': len(fin_sheets),
            'note_sheet_count': len(note_sheets),
            'rollover': do_rollover,
            'organize_notes': organize_notes,
            'period_change': period_change,
            'note_conflicts': {'duplicates': len(pre['note_conflicts']['duplicates']), 'missing_numbers': pre['note_conflicts']['missing_numbers'][:20]},
        }, ensure_ascii=True)
        return resp
    except Exception as e:
        return jsonify({'error': friendly_error_message(e, 'DSD→Excel 변환')}), 400

@app.route('/api/excel2dsd', methods=['POST'])
def api_excel2dsd():
    try:
        if 'xlsx' not in request.files:
            return jsonify({'error':'xlsx 파일이 필요합니다.'}), 400
        xlsx_bytes = request.files['xlsx'].read()
        chk = precheck_excel_bytes(xlsx_bytes)
        if not chk.get('ok'):
            return jsonify({'error':'재조립 전에 해결해야 할 치명 항목이 있습니다: ' + ' / '.join(chk.get('fatal',[]))}), 400
        dsd = excel_to_dsd_bytes(xlsx_bytes)
        resp = send_file(io.BytesIO(dsd), mimetype='application/octet-stream',
                         as_attachment=True, download_name='조립결과.dsd')
        resp.headers['X-Info'] = json.dumps({'status':'ok'}, ensure_ascii=True)
        return resp
    except Exception as e:
        return jsonify({'error': friendly_error_message(e, 'Excel→DSD 재조립')}), 400

@app.route('/api/ai_validate', methods=['POST'])
def api_ai_validate():
    try:
        if 'xlsx' not in request.files:
            return jsonify({'error':'xlsx 파일이 필요합니다.'}), 400
        provider = request.form.get('provider','gemini').strip()
        api_key = request.form.get('api_key','').strip()
        model = request.form.get('model','').strip()
        if not api_key:
            return jsonify({'error':'AI 검증은 API Key 입력 사용자만 사용할 수 있습니다.'}), 400
        summary = analyze_excel_financials(request.files['xlsx'].read())
        ai_text = run_ai_validation(summary, provider, api_key, model)
        return jsonify({
            'python_validation': summary,
            'ai_validation': ai_text,
            'provider': provider,
            'model': model
        })
    except Exception as e:
        return jsonify({'error': friendly_error_message(e, 'AI 검증')}), 400

@app.route('/api/verify_prior', methods=['POST'])
def api_verify_prior():
    try:
        if 'prev_dsd' not in request.files or 'curr_dsd' not in request.files:
            return jsonify({'error':'prev_dsd, curr_dsd 파일이 필요합니다.'}), 400
        return jsonify(verify_prior_period(request.files['prev_dsd'].read(), request.files['curr_dsd'].read()))
    except Exception as e:
        return jsonify({'error': friendly_error_message(e, '전기금액 검증')}), 400

@app.route('/api/diff_dsd', methods=['POST'])
def api_diff_dsd():
    try:
        if 'before_dsd' not in request.files or 'after_dsd' not in request.files:
            return jsonify({'error':'before_dsd, after_dsd 파일이 필요합니다.'}), 400
        return jsonify(compare_dsd_files(request.files['before_dsd'].read(), request.files['after_dsd'].read()))
    except Exception as e:
        return jsonify({'error': friendly_error_message(e, 'DSD 비교분석')}), 400

@app.route('/api/diff_dsd_xlsx', methods=['POST'])
def api_diff_dsd_xlsx():
    try:
        if 'before_dsd' not in request.files or 'after_dsd' not in request.files:
            return jsonify({'error':'before_dsd, after_dsd 파일이 필요합니다.'}), 400
        result = compare_dsd_files(request.files['before_dsd'].read(), request.files['after_dsd'].read())
        out = build_diff_report_xlsx(result)
        return send_file(io.BytesIO(out),
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True,
                         download_name='DSD_Diff_Report.xlsx')
    except Exception as e:
        return jsonify({'error': friendly_error_message(e, 'Diff 리포트 생성')}), 400

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    print(f"EasyDSD v0.13 서버 시작: http://localhost:{port}")
    app.run(host='0.0.0.0', port=port, debug=False)
